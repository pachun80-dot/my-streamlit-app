import io
import os
import re
import sys
import glob
import time
import threading
import unicodedata
import warnings
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment

# gRPC 경고 억제
os.environ['GRPC_ENABLE_FORK_SUPPORT'] = '0'
os.environ['GRPC_POLL_STRATEGY'] = 'poll'
warnings.filterwarnings('ignore', category=FutureWarning)

from pdf_parser import (
    parse_pdf, split_articles, _detect_lang,
    extract_structured_articles, save_structured_to_excel
)
from html_parser import parse_eu_html_to_dataframe, parse_china_html_to_dataframe, parse_nz_html_to_dataframe, parse_taiwan_html_to_dataframe, parse_germany_html_to_dataframe, parse_russia_html_to_dataframe
from translator import translate_batch, _clean_translation_output
from embedder import (
    find_similar_korean,
    find_similar_korean_ai,
    find_similar_korean_batch,
    select_relevant_korean_laws,
)

# ── 페이지 설정 ──────────────────────────────────────────────
st.set_page_config(
    page_title="법령 번역 비교 분석 시스템",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 데이터 경로 ──────────────────────────────────────────────
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

if sys.platform == "win32":
    DATA_DIR = os.environ.get("DATA_DIR", r"C:\Users\milhi\Desktop\DATA")
else:
    DATA_DIR = os.environ.get("DATA_DIR", os.path.join(PROJECT_DIR, "DATA"))

COUNTRY_MAP = {
    "남아프리카공화국": "SOUTH_AFRICA",
    "뉴질랜드": "NEWZEALAND",
    "대만": "TAIWAN",
    "독일": "GERMANY",
    "러시아": "RUSSIA",
    "말레이시아": "MALAYSIA",
    "미국": "USA",
    "베트남": "VIETNAM",
    "브라질": "BRAZIL",
    "사우디아라비아": "SAUDI_ARABIA",
    "싱가포르": "SINGAPORE",
    "아프리카지식재산권기구": "OAPI",
    "영국": "UK",
    "유럽(EPC)": "EPC",
    "유럽연합": "EU",
    "인도": "INDIA",
    "인도네시아": "INDONESIA",
    "일본": "JAPAN",
    "중국": "CHINA",
    "캐나다": "CANADA",
    "태국": "THAILAND",
    "튀르키에": "TURKIYE",
    "프랑스": "FRANCE",
    "한국": "KOREA",
    "호주": "AUSTRALIA",
    "홍콩": "HONGKONG",
}

KOREA_FOLDER = "KOREA"


def _safe_join(*parts: str) -> str:
    """한글 경로 호환성 처리 (macOS NFD vs Linux NFC).

    macOS에서 Git에 커밋된 한글 폴더명은 NFD(분해형)로 저장될 수 있다.
    Linux(Streamlit Cloud)에서는 NFC(조합형)와 NFD가 서로 다른 경로로 취급되므로,
    직접 구성한 경로가 존재하지 않을 때 부모 디렉토리를 스캔하여 일치하는 항목을 찾는다.
    """
    path = os.path.join(*parts)
    if os.path.exists(path):
        return path
    # NFC/NFD 변환 시도
    for form in ("NFC", "NFD"):
        normalized = unicodedata.normalize(form, path)
        if os.path.exists(normalized):
            return normalized
    # 부모 디렉토리에서 이름 매칭 시도
    parent = os.path.dirname(path)
    target = unicodedata.normalize("NFC", os.path.basename(path))
    if os.path.isdir(parent):
        for entry in os.listdir(parent):
            if unicodedata.normalize("NFC", entry) == target:
                return os.path.join(parent, entry)
    return path


# ── 전문 UI 스타일 ──────────────────────────────────────────────
PROFESSIONAL_STYLE = """
<style>
/* ── 테마 변수 (라이트 모드 기본) ── */
:root {
    --bg-primary: #fefcfb;
    --bg-secondary: #ffffff;
    --bg-tertiary: #fdf8f6;
    --bg-sidebar: linear-gradient(180deg, #fdf8f6 0%, #ffffff 100%);
    --text-primary: #5a4e53;
    --text-heading: #4a3d42;
    --text-secondary: #8a7e84;
    --border-color: #e8ddd8;
    --border-subtle: #f0e8e4;
    --accent: #8b2240;
    --accent-light: #a3324f;
    --accent-dark: #5c1a2a;
    --card-shadow: rgba(0, 0, 0, 0.04);
    --input-bg: #ffffff;
    --tab-hover-bg: #fdf4f0;
    --tab-active-bg: white;
    --table-even-bg: #fafafa;
    --badge-success-bg: #dcfce7;
    --badge-success-text: #166534;
    --badge-warning-bg: #fef3c7;
    --badge-warning-text: #92400e;
    --badge-info-bg: #fce8ed;
    --badge-info-text: #7a1b33;
    --diff-bg: #fff8e1;
    --diff-border: #ffc107;
    --korea-bg: #e8f5e9;
    --korea-border: #4caf50;
    --alert-bg: #ffffff;
}

/* ── 다크 모드 ── */
@media (prefers-color-scheme: dark) {
  :root {
    --bg-primary: #1a1a2e;
    --bg-secondary: #252540;
    --bg-tertiary: #2d2d48;
    --bg-sidebar: linear-gradient(180deg, #1e1e35 0%, #252540 100%);
    --text-primary: #ddd6d9;
    --text-heading: #ede7ea;
    --text-secondary: #a89da3;
    --border-color: #3e3e58;
    --border-subtle: #33334a;
    --accent: #d4587a;
    --accent-light: #e06b8a;
    --accent-dark: #a3324f;
    --card-shadow: rgba(0, 0, 0, 0.3);
    --input-bg: #2d2d48;
    --tab-hover-bg: #33334a;
    --tab-active-bg: #2d2d48;
    --table-even-bg: #2a2a44;
    --badge-success-bg: #1a3a2a;
    --badge-success-text: #86efac;
    --badge-warning-bg: #3d2e10;
    --badge-warning-text: #fcd34d;
    --badge-info-bg: #3a1a28;
    --badge-info-text: #f9a8c0;
    --diff-bg: #2e2a1a;
    --diff-border: #b8960e;
    --korea-bg: #1a2e1e;
    --korea-border: #388e3c;
    --alert-bg: #252540;
  }
}

/* 전체 폰트 및 테마 적용 */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

.stApp {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    background-color: var(--bg-primary) !important;
    color: var(--text-primary) !important;
}

/* 메인 콘텐츠 영역 */
.stApp > header {
    background-color: var(--bg-primary) !important;
}

.main .block-container {
    background-color: var(--bg-primary) !important;
    color: var(--text-primary) !important;
}

/* 모든 텍스트 요소 기본 색상 */
.stApp p, .stApp span, .stApp label, .stApp div,
.stApp li, .stApp td, .stApp th, .stApp caption {
    color: var(--text-primary) !important;
}

.stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {
    color: var(--text-heading) !important;
}

/* 메인 헤더는 흰색 텍스트 유지 (크림슨 배경) */
.stApp .main-header,
.stApp .main-header h1,
.stApp .main-header p,
.stApp .main-header span,
.stApp .main-header div {
    color: #ffffff !important;
}

/* 사이드바 스타일 */
section[data-testid="stSidebar"] {
    background: var(--bg-sidebar) !important;
    border-right: 1px solid var(--border-color);
}

section[data-testid="stSidebar"] .block-container {
    padding-top: 2rem;
}

section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] div {
    color: var(--text-primary) !important;
}

/* 라디오 버튼 / 체크박스 라벨 */
.stRadio label, .stCheckbox label {
    color: var(--text-primary) !important;
}

/* 셀렉트박스 / 멀티셀렉트 텍스트 */
[data-baseweb="select"] span,
[data-baseweb="select"] div {
    color: var(--text-primary) !important;
}

/* 멀티셀렉트 태그 (선택된 항목) */
[data-baseweb="tag"] {
    color: var(--text-primary) !important;
    background-color: var(--bg-tertiary) !important;
}
[data-baseweb="tag"] span {
    color: var(--text-primary) !important;
}

/* 드롭다운 메뉴 / 팝오버 */
[data-baseweb="popover"] {
    background-color: var(--bg-secondary) !important;
}
[data-baseweb="menu"],
[data-baseweb="popover"] ul {
    background-color: var(--bg-secondary) !important;
}
[data-baseweb="menu"] li,
[data-baseweb="menu"] li span,
[data-baseweb="menu"] li div {
    color: var(--text-primary) !important;
    background-color: var(--bg-secondary) !important;
}
[data-baseweb="menu"] li:hover,
[data-baseweb="menu"] li:hover span {
    background-color: var(--bg-tertiary) !important;
}

/* caption 텍스트 */
.stApp .stCaption, .stApp small,
.stApp [data-testid="stCaptionContainer"] p {
    color: var(--text-secondary) !important;
}

/* 체크박스 내부 텍스트 */
.stApp .stCheckbox span,
.stApp .stCheckbox label,
.stApp .stCheckbox div {
    color: var(--text-primary) !important;
}

/* divider */
.stApp hr {
    border-color: var(--border-color) !important;
}

/* warning/info/error 메시지 내부 텍스트 */
.stApp .stAlert p,
.stApp .stAlert span,
.stApp .stAlert div {
    color: inherit !important;
}

/* 메인 헤더 */
.main-header {
    background: linear-gradient(135deg, var(--accent-dark) 0%, var(--accent) 50%, var(--accent-light) 100%);
    color: white;
    padding: 2rem 2.5rem;
    border-radius: 12px;
    margin-bottom: 2rem;
    box-shadow: 0 4px 12px -2px rgba(92, 26, 42, 0.25);
}

.main-header h1 {
    margin: 0;
    font-size: 2rem;
    font-weight: 700;
    letter-spacing: -0.02em;
}

.main-header p {
    margin: 0.5rem 0 0 0;
    font-size: 1rem;
    opacity: 0.92;
}

/* 카드 스타일 */
.info-card {
    background: var(--bg-secondary) !important;
    padding: 1.5rem;
    border-radius: 10px;
    border: 1px solid var(--border-color);
    box-shadow: 0 1px 3px 0 var(--card-shadow);
    margin-bottom: 1rem;
}

.info-card h3 {
    color: var(--text-heading) !important;
    font-size: 1.1rem;
    font-weight: 600;
    margin-bottom: 0.5rem;
}

.info-card p {
    color: var(--text-secondary) !important;
}

/* 상태 배지 */
.status-badge {
    display: inline-block;
    padding: 0.25rem 0.75rem;
    border-radius: 9999px;
    font-size: 0.875rem;
    font-weight: 500;
    margin: 0.25rem;
}

.status-success {
    background: var(--badge-success-bg) !important;
    color: var(--badge-success-text) !important;
}

.status-warning {
    background: var(--badge-warning-bg) !important;
    color: var(--badge-warning-text) !important;
}

.status-info {
    background: var(--badge-info-bg) !important;
    color: var(--badge-info-text) !important;
}

/* 버튼 개선 (크림슨 배경) */
.stApp .stButton > button {
    background: linear-gradient(135deg, var(--accent) 0%, #6e1a33 100%) !important;
    color: #ffffff !important;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 1.5rem;
    font-weight: 600;
    font-size: 0.95rem;
    transition: all 0.2s;
    box-shadow: 0 2px 6px rgba(139, 34, 64, 0.2);
}
.stApp .stButton > button span,
.stApp .stButton > button div,
.stApp .stButton > button p { color: #ffffff !important; }

.stApp .stButton > button:hover {
    background: linear-gradient(135deg, #6e1a33 0%, var(--accent-dark) 100%) !important;
    color: #ffffff !important;
    box-shadow: 0 4px 10px rgba(139, 34, 64, 0.3);
    transform: translateY(-1px);
}

/* 다운로드 버튼 (초록 배경) */
.stApp .stDownloadButton > button {
    background: linear-gradient(135deg, #4a6741 0%, #3b5534 100%) !important;
    color: #ffffff !important;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 1.5rem;
    font-weight: 600;
    font-size: 0.95rem;
    box-shadow: 0 2px 4px rgba(74, 103, 65, 0.2);
}
.stApp .stDownloadButton > button span,
.stApp .stDownloadButton > button div,
.stApp .stDownloadButton > button p { color: #ffffff !important; }

.stApp .stDownloadButton > button:hover {
    background: linear-gradient(135deg, #3b5534 0%, #2d422a 100%) !important;
    color: #ffffff !important;
    box-shadow: 0 4px 8px rgba(74, 103, 65, 0.3);
    transform: translateY(-1px);
}

/* 입력 필드 */
.stSelectbox, .stMultiselect, .stTextInput {
    border-radius: 8px;
}

.stSelectbox > div > div, .stMultiselect > div > div, .stTextInput > div > div {
    border-radius: 8px;
    border-color: var(--border-color);
    background-color: var(--input-bg) !important;
    color: var(--text-primary) !important;
}

.stSelectbox > div > div:focus-within, .stMultiselect > div > div:focus-within {
    border-color: var(--accent);
    box-shadow: 0 0 0 1px var(--accent);
}

/* 텍스트 입력 */
.stTextInput input, .stTextArea textarea {
    background-color: var(--input-bg) !important;
    color: var(--text-primary) !important;
}

/* 메트릭 카드 */
.stMetric {
    background: var(--bg-secondary) !important;
    padding: 1rem;
    border-radius: 10px;
    border: 1px solid var(--border-color);
    box-shadow: 0 1px 2px 0 var(--card-shadow);
}

.stMetric label, .stMetric [data-testid="stMetricValue"],
.stMetric [data-testid="stMetricLabel"] {
    color: var(--text-primary) !important;
}

/* 탭 스타일 */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
    background: transparent;
    border-bottom: 2px solid var(--border-color);
}

.stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 8px 8px 0 0;
    padding: 0.75rem 1.5rem;
    font-weight: 500;
    color: var(--text-secondary) !important;
}

.stTabs [data-baseweb="tab"]:hover {
    background: var(--tab-hover-bg);
    color: var(--text-heading) !important;
}

.stTabs [aria-selected="true"] {
    background: var(--tab-active-bg);
    color: var(--accent) !important;
    border-bottom: 3px solid var(--accent);
}

/* 프로그레스 바 - 트랙 (배경) */
.stProgress > div > div > div {
    background-color: var(--border-color) !important;
    border-radius: 9999px;
}
/* 프로그레스 바 - 채움 (진행) */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, var(--accent) 0%, var(--accent-light) 100%) !important;
    border-radius: 9999px;
}

/* 알림 박스 */
.stAlert {
    border-radius: 10px;
    border-left: 4px solid;
    background-color: var(--alert-bg) !important;
}

/* 데이터프레임 테이블 */
.dataframe {
    border: 1px solid var(--border-color) !important;
    border-radius: 8px;
    overflow: hidden;
}

.dataframe thead tr th {
    background: var(--bg-tertiary) !important;
    color: var(--text-heading) !important;
    font-weight: 600 !important;
    border-bottom: 2px solid var(--border-color) !important;
}

.dataframe tbody tr td {
    background: var(--bg-secondary) !important;
    color: var(--text-primary) !important;
}

/* 사이드바 로고 영역 */
.sidebar-logo {
    text-align: center;
    padding: 1.5rem 1rem 1.5rem 1rem;
    margin-bottom: 1.5rem;
    background: linear-gradient(135deg, var(--accent-dark) 0%, var(--accent) 100%);
    border-radius: 10px;
    margin: 0 0.5rem 1.5rem 0.5rem;
}

section[data-testid="stSidebar"] .sidebar-logo h2,
.stApp .sidebar-logo h2 {
    color: #ffffff !important;
    font-size: 1.4rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: 0.02em;
}

section[data-testid="stSidebar"] .sidebar-logo p,
.stApp .sidebar-logo p {
    color: rgba(255,255,255,0.85) !important;
    font-size: 0.82rem;
    margin: 0.3rem 0 0 0;
    font-weight: 400;
}

/* 섹션 헤더 */
.section-header {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0 0 0.75rem 0;
    border-bottom: 3px solid var(--accent);
    margin-bottom: 1.5rem;
}

.section-header h3 {
    color: var(--text-heading) !important;
    font-size: 1.3rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: -0.01em;
}

/* expander */
.streamlit-expanderHeader {
    background-color: var(--bg-secondary) !important;
    color: var(--text-primary) !important;
}

.streamlit-expanderContent {
    background-color: var(--bg-secondary) !important;
    color: var(--text-primary) !important;
}

/* status 위젯 */
[data-testid="stStatusWidget"] {
    background-color: var(--bg-secondary) !important;
    color: var(--text-primary) !important;
}
</style>
"""

st.markdown(PROFESSIONAL_STYLE, unsafe_allow_html=True)

# ── 메인 헤더 ──────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>법령 번역 비교 분석 시스템</h1>
    <p>다국어 특허법을 AI 번역하고 한국법과 비교 분석하는 전문 플랫폼</p>
</div>
""", unsafe_allow_html=True)


# ── 유틸리티 함수 ────────────────────────────────────────────

def _list_pdfs(folder: str) -> list[str]:
    """PDF, XML, RTF 파일 목록을 반환한다."""
    path = os.path.join(DATA_DIR, folder)
    if not os.path.isdir(path):
        return []
    pdfs = glob.glob(os.path.join(path, "*.pdf"))
    xmls = glob.glob(os.path.join(path, "*.xml"))
    rtfs = glob.glob(os.path.join(path, "*.rtf"))
    return sorted(pdfs + xmls + rtfs)


def _list_result_files() -> list[str]:
    """번역 결과 Excel 파일 목록을 반환한다."""
    files = []
    translation_dir = _safe_join(DATA_DIR, "output", "번역비교결과")

    # 국가별 하위 폴더에서 검색 (NFC/NFD 호환 방식)
    if os.path.isdir(translation_dir):
        for entry in sorted(os.listdir(translation_dir)):
            country_dir = _safe_join(translation_dir, entry)
            if os.path.isdir(country_dir):
                files.extend(_safe_glob(country_dir, "*.xlsx"))

        # 하위 호환성: 번역비교결과 폴더 루트의 파일도 포함
        files.extend(_safe_glob(translation_dir, "*.xlsx"))

    # 하위 호환성: 기존 output 폴더의 번역 파일도 포함
    output_dir = _safe_join(DATA_DIR, "output")
    if os.path.isdir(output_dir):
        files.extend(_safe_glob(output_dir, "번역비교_*.xlsx"))

    # 중복 제거 및 최신순 정렬
    files = list(dict.fromkeys(files))
    def _safe_mtime(f):
        try:
            return os.path.getmtime(f)
        except Exception:
            return 0
    return sorted([f for f in files if not _basename(f).startswith("~$")],
                  key=_safe_mtime, reverse=True)


def _safe_glob(directory: str, pattern: str) -> list[str]:
    """한글 파일명 호환 glob. NFC/NFD 정규화 문제를 우회한다."""
    results = glob.glob(os.path.join(directory, pattern))
    if results:
        return results
    # glob 실패 시 os.listdir + fnmatch로 NFC 비교
    import fnmatch
    try:
        entries = os.listdir(directory)
    except OSError:
        return []
    nfc_pattern = unicodedata.normalize("NFC", pattern)
    matched = []
    for entry in entries:
        nfc_entry = unicodedata.normalize("NFC", entry)
        if fnmatch.fnmatch(nfc_entry, nfc_pattern):
            matched.append(os.path.join(directory, entry))
    return matched


def _list_structured_excels() -> list[str]:
    """구조화법률 폴더 내 구조화 엑셀 파일 목록 (한국법 제외)."""
    structured_dir = _safe_join(DATA_DIR, "output", "구조화법률")
    files = []

    # 국가별 하위 폴더에서 검색 (한국 제외)
    if os.path.isdir(structured_dir):
        for country in COUNTRY_MAP.keys():
            if country == '한국':
                continue
            country_dir = os.path.join(structured_dir, country)
            if os.path.isdir(country_dir):
                files.extend(glob.glob(os.path.join(country_dir, "*.xlsx")))

        # 하위 호환성: 구조화법률 폴더 루트의 파일도 포함
        files.extend(_safe_glob(structured_dir, "구조화_*.xlsx"))
        files.extend(_safe_glob(structured_dir, "EU_*.xlsx"))

    # 하위 호환성: 기존 output 폴더의 구조화 파일도 포함
    output_dir = _safe_join(DATA_DIR, "output")
    if os.path.isdir(output_dir):
        files.extend(_safe_glob(output_dir, "구조화_*.xlsx"))

    all_files = sorted(
        [f for f in files if not _basename(f).startswith("~$")],
        key=os.path.getmtime, reverse=True,
    )
    # 한국법 구조화 엑셀 제외
    nfc = unicodedata.normalize
    return [f for f in all_files
            if "한국" not in nfc("NFC", _basename(f)) and "KOREA" not in _basename(f).upper()]


def _list_korea_excels() -> list[str]:
    """구조화법률 폴더 내 한국법 구조화 엑셀 파일 목록."""
    structured_dir = _safe_join(DATA_DIR, "output", "구조화법률")
    files = []

    # 한국 하위 폴더에서 검색
    if os.path.isdir(structured_dir):
        korea_dir = os.path.join(structured_dir, "한국")
        if os.path.isdir(korea_dir):
            files.extend(glob.glob(os.path.join(korea_dir, "*.xlsx")))

        # 하위 호환성: 구조화법률 폴더 루트의 한국법 파일도 포함
        files.extend(_safe_glob(structured_dir, "구조화_한국_*.xlsx"))

    # 하위 호환성: 기존 output 폴더도 확인
    output_dir = _safe_join(DATA_DIR, "output")
    if os.path.isdir(output_dir):
        files.extend(_safe_glob(output_dir, "구조화_한국_*.xlsx"))

    return sorted(
        [f for f in files if not _basename(f).startswith("~$")],
        key=os.path.getmtime, reverse=True,
    )


def _basename(path: str) -> str:
    return os.path.basename(path)


def _detect_country_from_filename(filename: str) -> str:
    """파일명에서 국가를 감지한다.

    Returns:
        국가명 또는 빈 문자열
    """
    filename_lower = filename.lower()

    # 파일명 패턴 매칭 (구체적인 키워드 → 일반적인 키워드 순서)
    _COUNTRY_KEYWORDS = [
        ('일본', ['japan', '일본', '334ac']),
        ('중국', ['china', '중국', 'cnipa']),
        ('미국', ['usa', 'united states', '미국', 'title35', 'westlaw']),
        ('유럽(EPC)', ['epc', 'european patent']),
        ('유럽연합', ['eu_', '유럽연합']),
        ('독일', ['germany', '독일', 'bjnr']),
        ('영국', ['uk', 'united kingdom', '영국']),
        ('프랑스', ['france', '프랑스']),
        ('인도네시아', ['indonesia', '인도네시아']),
        ('인도', ['india', '인도']),
        ('베트남', ['vietnam', '베트남']),
        ('브라질', ['brazil', '브라질']),
        ('러시아', ['russia', '러시아']),
        ('싱가포르', ['singapore', '싱가포르']),
        ('호주', ['australia', '호주']),
        ('대만', ['taiwan', '대만']),
        ('태국', ['thailand', '태국']),
        ('아프리카지식재산권기구', ['oapi', '아프리카지식재산권']),
        ('남아프리카공화국', ['south africa', '남아프리카']),
        ('사우디아라비아', ['saudi', '사우디']),
        ('캐나다', ['canada', '캐나다']),
        ('말레이시아', ['malaysia', '말레이시아']),
        ('튀르키에', ['turkiye', 'turkey', '튀르키에']),
        ('홍콩', ['hongkong', 'hong kong', '홍콩', 'cap ']),
        ('뉴질랜드', ['newzealand', 'new zealand', '뉴질랜드']),
        ('한국', ['korea', '한국']),
    ]

    for country, keywords in _COUNTRY_KEYWORDS:
        for kw in keywords:
            if kw in filename or kw in filename_lower:
                return country

    return ''


def _country_to_folder_name(country: str) -> str:
    """국가명을 폴더명으로 변환한다.

    Args:
        country: _detect_country_from_filename()이 반환한 국가명

    Returns:
        실제 폴더명
    """
    # 특수 매핑 (괄호 제거 등)
    _FOLDER_MAPPING = {
        '유럽(EPC)': '유럽',
    }

    return _FOLDER_MAPPING.get(country, country)


def _korean_law_name(source: str) -> str:
    """PDF/Excel 파일명에서 한국법 명칭 추출. 예: '구조화_한국_특허법(법률)(...).xlsx' → '한국_특허법'"""
    name = source.replace(".pdf", "").replace(".PDF", "").replace(".xlsx", "").replace(".XLSX", "").replace(".rtf", "").replace(".RTF", "")
    if "(" in name:
        name = name[:name.index("(")]
    name = name.strip()

    # "구조화_한국_" 접두사 제거
    if name.startswith("구조화_한국_"):
        name = name.replace("구조화_한국_", "한국_", 1)
    elif name.startswith("구조화_"):
        name = name.replace("구조화_", "", 1)

    # 이미 "한국"으로 시작하지 않으면 "한국_" 추가
    if not name.startswith("한국"):
        name = f"한국_{name}"

    return name if name else "한국법"


def _article_num_display(article_num: str) -> str:
    """조문번호 UI 표시용 변환: 'Section N' / 'Article N' / '§ N' → 'N' (내부 데이터는 변경하지 않음)."""
    s = str(article_num).strip()
    m = re.match(r'^(?:Section|Article|Rule|§)\s*(.+)$', s, re.IGNORECASE)
    return m.group(1).strip() if m else s


def _clean_text(text: str) -> str:
    """법률 조문과 관련 없는 텍스트와 마크다운 기호를 제거한다."""
    if not text or not isinstance(text, str):
        return ""
    s = text.strip()
    # 마크다운 기호 제거
    s = re.sub(r"[*_#`~>|]", "", s)
    s = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", s)  # [text](url) → text
    # 페이지 번호 / 머리글 / 바닥글 패턴 제거
    s = re.sub(r"(?m)^[-─━=]{3,}$", "", s)
    s = re.sub(r"(?m)^Page\s*\d+.*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"(?m)^-\s*\d+\s*-\s*$", "", s)
    s = re.sub(r"(?m)^列印時間[：:].*$", "", s)
    s = re.sub(r"(?m)^所有條文\s*$", "", s)
    s = re.sub(r"(?m)^法規名稱[：:].*$", "", s)
    s = re.sub(r"(?m)^修正日期[：:].*$", "", s)
    s = re.sub(r"(?m)^修正⽇期[：:].*$", "", s)
    s = re.sub(r"(?m)^法規類別[：:].*$", "", s)
    # 연속 빈 줄 정리
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _esc(text: str) -> str:
    """HTML 이스케이프."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ── 공통 스타일 ──────────────────────────────────────────────
DETAIL_STYLE = """
<style>
.article-container {
    display: flex;
    margin-bottom: 20px;
    gap: 16px;
}
.article-structure {
    flex: 0 0 280px;
    background: var(--bg-tertiary) !important;
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 14px;
}
.structure-title {
    font-weight: 700;
    font-size: 0.95em;
    color: var(--text-heading) !important;
    margin-bottom: 8px;
    padding-bottom: 6px;
    border-bottom: 1px solid var(--border-color);
    background-color: transparent !important;
}
.structure-content {
    font-size: 0.88em;
    line-height: 1.8;
    color: var(--text-primary) !important;
    background-color: transparent !important;
    white-space: pre-wrap;
    word-break: break-word;
}
.article-row {
    display: flex;
    border: 1px solid var(--border-color);
    border-radius: 8px;
    overflow: hidden;
    flex: 1;
    background-color: var(--bg-secondary) !important;
}
.article-col {
    flex: 1;
    padding: 12px 16px;
    border-right: 2px solid var(--border-color);
    min-width: 0;
    background-color: var(--bg-secondary) !important;
    color: var(--text-primary) !important;
}
.article-col:last-child { border-right: none; }
.article-col-header {
    font-weight: 700; font-size: 0.95em;
    margin-bottom: 8px; padding-bottom: 6px;
    border-bottom: 1px solid var(--border-color);
}
.col-original .article-col-header { color: var(--accent) !important; }
.col-gemini .article-col-header { color: #b8860b !important; }
.col-claude .article-col-header { color: var(--accent-dark) !important; }
.article-col-body {
    font-size: 0.9em; line-height: 1.6;
    white-space: pre-wrap; word-break: break-word;
    color: var(--text-primary) !important; background-color: var(--bg-secondary) !important;
}
.diff-box {
    background: var(--diff-bg) !important; border-left: 4px solid var(--diff-border);
    padding: 10px 14px; margin: 8px 0;
    border-radius: 0 6px 6px 0; font-size: 0.9em; line-height: 1.6;
    color: var(--text-primary) !important;
}
.diff-box strong { color: var(--text-heading) !important; }
.korea-law-box {
    background: var(--korea-bg) !important; border-left: 4px solid var(--korea-border);
    padding: 10px 14px; margin: 8px 0;
    border-radius: 0 6px 6px 0; font-size: 0.9em; line-height: 1.6;
    color: var(--text-primary) !important;
}
.korea-law-box strong { color: var(--text-heading) !important; }
.article-title { font-size: 1.1em; font-weight: 700; margin-bottom: 6px; color: var(--text-heading) !important; }
/* 전체 보기 테이블 */
.fullview-table { width: 100%; border-collapse: collapse; table-layout: fixed; }
.fullview-table th {
    background: var(--bg-tertiary) !important; padding: 10px 12px;
    border: 1px solid var(--border-color); text-align: left;
    position: sticky; top: 0; z-index: 1;
    color: var(--text-heading) !important;
}
.fullview-table td {
    padding: 10px 12px; border: 1px solid var(--border-color);
    vertical-align: top; white-space: pre-wrap;
    word-break: break-word; font-size: 0.88em; line-height: 1.6;
    color: var(--text-primary) !important; background-color: var(--bg-secondary) !important;
}
.fullview-table tr:nth-child(even) { background: var(--table-even-bg) !important; }
.fullview-table tr:nth-child(even) td { background: var(--table-even-bg) !important; }
.fullview-table col.col-id { width: 10% !important; max-width: 10% !important; }
.fullview-table col.col-text { width: 30% !important; max-width: 30% !important; }
.fullview-table col.col-text-narrow { width: 22.5% !important; max-width: 22.5% !important; }
.fullview-table col.col-korean { width: 22.5% !important; max-width: 22.5% !important; }
</style>
"""

# ── 사이드바 네비게이션 ────────────────────────────────────────
with st.sidebar:
    # 사이드바 로고
    st.markdown("""
    <div class="sidebar-logo">
        <h2>LegalAI</h2>
        <p>법령 번역 분석 플랫폼</p>
    </div>
    """, unsafe_allow_html=True)

    # API 키 상태
    st.markdown("### 시스템 상태")

    try:
        gemini_api = st.secrets.get("GOOGLE_API_KEY", "")
    except Exception:
        gemini_api = ""
    try:
        claude_api = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        claude_api = ""

    gemini_status = "success" if gemini_api else "warning"
    claude_status = "success" if claude_api else "warning"

    st.markdown(f"""
    <div style="margin-bottom: 1rem;">
        <span class="status-badge status-{gemini_status}">
            {'Connected' if gemini_api else 'Not set'} Gemini API
        </span>
        <br>
        <span class="status-badge status-{claude_status}">
            {'Connected' if claude_api else 'Not set'} Claude API
        </span>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # 네비게이션
    st.markdown("### 메뉴")
    page = st.radio(
        "기능 선택",
        ["법령 구조화", "번역 실행", "상세보기"],
        label_visibility="collapsed"
    )

    st.divider()

    # 도움말
    with st.expander("사용 가이드"):
        st.markdown("""
        **1단계: 법령 구조화**
        - PDF/XML 파일에서 조문 자동 추출

        **2단계: 번역 실행**
        - Gemini / Claude 이중 번역
        - 한국법 유사 조문 매칭

        **3단계: 상세보기**
        - 번역 비교 분석 및 결과 다운로드
        """)

# ══════════════════════════════════════════════════════════════
# 페이지 1: 법령 구조화 (PDF → 구조화 엑셀)
# ══════════════════════════════════════════════════════════════
if page == "법령 구조화":
    st.markdown("""
    <div class="section-header">
        <h3>법령 구조화</h3>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <p style="margin: 0; color: #64748b;">
            법령 PDF/XML/HTML 파일을 자동으로 분석하여 편/장/절/조/항/호 단위로 구조화합니다.
            구조화된 데이터는 Excel 파일로 저장되어 번역 작업에 활용됩니다.
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Step 1: 국가 선택
    struct_country = st.selectbox(
        "📍 국가 선택",
        list(COUNTRY_MAP.keys()),
        key="struct_country",
        help="먼저 국가를 선택하세요"
    )

    # Step 2: 국가에 따라 입력 방식 표시
    if struct_country == "일본":
        # 일본은 파일 업로드만 지원
        st.info("💡 일본 법령은 다운로드한 HTML 파일을 업로드하세요")

        uploaded_file = st.file_uploader(
            "일본 법령 HTML 파일 업로드",
            type=['html'],
            help="e-Gov 법령検索에서 다운로드한 HTML 파일",
            key="japan_html_upload"
        )

        input_method = "파일 업로드"
        html_url = None
        struct_pdf_selected = None

    elif struct_country == "프랑스":
        # 프랑스는 LEGI XML 디렉토리 경로 입력
        st.info("💡 프랑스 법령은 LEGI XML 디렉토리 경로를 입력하세요")

        input_method = st.radio(
            "📥 입력 방식",
            ["디렉토리 경로 입력", "DATA 폴더에서 선택"],
            horizontal=True,
            key="france_input_method"
        )

        if input_method == "디렉토리 경로 입력":
            legi_dir = st.text_input(
                "LEGI XML 디렉토리 경로",
                value="DATA/FRANCE/CPI_only/LEGITEXT000006069414",
                help="article/LEGI/ARTI 폴더가 포함된 디렉토리 경로",
                key="france_legi_dir"
            )
            html_url = legi_dir  # 경로를 html_url 변수에 저장
            uploaded_file = None
            struct_pdf_selected = None
        else:
            # DATA/FRANCE 폴더에서 선택
            france_folder = os.path.join(DATA_DIR, "FRANCE")
            legi_dirs = []
            if os.path.exists(france_folder):
                for root, dirs, files in os.walk(france_folder):
                    # LEGITEXT로 시작하는 디렉토리 찾기
                    for d in dirs:
                        if d.startswith("LEGITEXT"):
                            full_path = os.path.join(root, d)
                            # article 폴더가 있는지 확인
                            if os.path.exists(os.path.join(full_path, "article")):
                                rel_path = os.path.relpath(full_path, DATA_DIR)
                                legi_dirs.append(rel_path)

            if not legi_dirs:
                st.warning(f"`{france_folder}/` 폴더에 LEGI XML 디렉토리를 넣어주세요.")

            selected_legi = st.selectbox(
                "LEGI 디렉토리 선택",
                legi_dirs if legi_dirs else [""],
                disabled=not legi_dirs,
                key="france_legi_select"
            )

            html_url = os.path.join(DATA_DIR, selected_legi) if selected_legi else None
            uploaded_file = None
            struct_pdf_selected = None
    
    elif struct_country in ["중국", "뉴질랜드", "대만", "독일", "러시아"]:
        # HTML URL 전용 (PDF 지원 안 함)
        st.info("💡 해당 국가는 공식 법령 웹사이트 URL을 입력하세요")

        input_method = "HTML URL 입력"

        _url_placeholders = {
            "중국": "https://www.cnipa.gov.cn/art/2020/11/23/art_97_155167.html",
            "뉴질랜드": "https://www.legislation.govt.nz/act/public/2013/0068/latest/DLM1419624.html",
            "대만": "https://law.moj.gov.tw/ENG/LawClass/LawAll.aspx?pcode=J0070007",
            "독일": "https://www.gesetze-im-internet.de/patg/BJNR201170936.html",
            "러시아": "https://rospatent.gov.ru/en/documents/grazhdanskiy-kodeks-rossiyskoy-federacii-chast-chetvertaya",
        }
        html_url = st.text_input(
            "법령 HTML URL",
            placeholder=_url_placeholders.get(struct_country, ""),
            key="html_url"
        )
        uploaded_file = None
        struct_pdf_selected = None

    elif struct_country == "유럽(EPC)":
        # 유럽은 HTML URL 또는 PDF 파일
        input_method = st.radio(
            "📥 입력 방식",
            ["HTML URL 입력", "파일 업로드"],
            horizontal=True,
            key="input_method"
        )

        if input_method == "HTML URL 입력":
            html_url = st.text_input(
                "법령 HTML URL",
                placeholder="https://eur-lex.europa.eu/...",
                key="html_url"
            )
            uploaded_file = None
            struct_pdf_selected = None
        else:
            html_url = None
            struct_folder = COUNTRY_MAP[struct_country]
            struct_pdfs = _list_pdfs(struct_folder)

            if not struct_pdfs:
                st.warning(f"`{DATA_DIR}/{struct_folder}/` 폴더에 파일을 넣어주세요.")

            struct_pdf_selected = st.selectbox(
                "법령 파일 선택",
                struct_pdfs,
                format_func=_basename,
                disabled=not struct_pdfs,
                key="struct_pdf",
            )
            uploaded_file = None
    else:
        # 기타 국가는 파일 업로드만
        input_method = "파일 업로드"
        html_url = None
        uploaded_file = None

        struct_folder = COUNTRY_MAP[struct_country]
        struct_pdfs = _list_pdfs(struct_folder)

        if not struct_pdfs:
            st.warning(f"`{DATA_DIR}/{struct_folder}/` 폴더에 PDF 또는 XML 파일을 넣어주세요.")

        struct_pdf_selected = st.selectbox(
            "법령 파일 선택",
            struct_pdfs,
            format_func=_basename,
            disabled=not struct_pdfs,
            key="struct_pdf",
        )

    # 실행 버튼
    can_run = False
    if struct_country == "일본":
        can_run = uploaded_file is not None
    elif struct_country == "프랑스":
        can_run = html_url is not None and html_url.strip() != ""
    elif input_method == "HTML URL 입력" or input_method == "디렉토리 경로 입력":
        can_run = html_url is not None and html_url.strip() != ""
    else:
        can_run = struct_pdf_selected is not None

    struct_run = st.button(
        "🚀 구조화 실행",
        type="primary",
        disabled=not can_run,
        key="struct_run",
    )

    if struct_run:
        output_dir = os.path.join(DATA_DIR, "output")
        os.makedirs(output_dir, exist_ok=True)

        with st.status("법령 구조화 파싱 중...", expanded=True) as status:
            # 일본 HTML 파일 업로드 처리
            if struct_country == "일본" and uploaded_file:
                st.write(f"일본 법령 HTML 파싱 중: {uploaded_file.name}")
                try:
                    # 임시 파일로 저장
                    import tempfile
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.html') as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        tmp_path = tmp_file.name

                    # 일본 파서 사용
                    from japan_parser import parse_japan_html_to_dataframe
                    df_structured = parse_japan_html_to_dataframe(tmp_path)

                    st.write(f"{len(df_structured)}개 항목 추출 (章/節/條/項/號 단위)")

                    # 임시 파일 삭제
                    os.unlink(tmp_path)

                    # 파일명 생성
                    base_name = os.path.splitext(uploaded_file.name)[0]
                except Exception as e:
                    st.error(f"일본 법령 파싱 실패: {str(e)}")
                    import traceback
                    st.error(traceback.format_exc())
                    st.stop()

            # 프랑스 LEGI XML 디렉토리 처리
            elif struct_country == "프랑스" and html_url:
                st.write(f"프랑스 LEGI XML 파싱 중: {html_url}")
                try:
                    from parsers.france import parse_french_legi_xml
                    import pandas as pd

                    # L조문 파싱
                    st.write("📖 L조문 (Partie législative) 파싱 중...")
                    df_l = parse_french_legi_xml(html_url, "L")
                    l_count = len(df_l)
                    st.write(f"  ✓ {l_count}개 행 추출")

                    # R조문 파싱
                    st.write("📖 R조문 (Partie réglementaire) 파싱 중...")
                    df_r = parse_french_legi_xml(html_url, "R")
                    r_count = len(df_r)
                    st.write(f"  ✓ {r_count}개 행 추출")

                    # 전체 합치기
                    df_structured = pd.concat([df_l, df_r], ignore_index=True)
                    st.write(f"✅ 총 {len(df_structured)}개 항목 추출 (편/장/절/조/항/호/목 단위)")

                    # 파일명 생성
                    law_name = os.path.basename(html_url)
                    base_name = f"프랑스_{law_name}"
                except Exception as e:
                    st.error(f"프랑스 LEGI XML 파싱 실패: {str(e)}")
                    import traceback
                    st.error(traceback.format_exc())
                    st.stop()

            # HTML URL 입력 방식
            elif input_method == "HTML URL 입력":
                st.write(f"HTML 파싱 중: {html_url}")
                try:
                    # 국가별 파서 선택
                    if struct_country == "중국":
                        df_structured = parse_china_html_to_dataframe(html_url)
                    elif struct_country == "뉴질랜드":
                        df_structured = parse_nz_html_to_dataframe(html_url)
                    elif struct_country == "대만":
                        df_structured = parse_taiwan_html_to_dataframe(html_url)
                    elif struct_country == "독일":
                        df_structured = parse_germany_html_to_dataframe(html_url)
                    elif struct_country == "러시아":
                        df_structured = parse_russia_html_to_dataframe(html_url)
                    else:
                        df_structured = parse_eu_html_to_dataframe(html_url)

                    st.write(f"{len(df_structured)}개 항목 추출 (조/항/호 단위)")

                    # 파일명 생성 (URL에서 추출)
                    import hashlib
                    url_hash = hashlib.md5(html_url.encode()).hexdigest()[:8]
                    base_name = f"{struct_country}_HTML_{url_hash}"
                except Exception as e:
                    st.error(f"HTML 파싱 실패: {str(e)}")
                    st.stop()

            # 파일 선택 방식
            elif struct_pdf_selected:
                st.write(f"{_basename(struct_pdf_selected)} 처리 중...")

                # XML 파일인지 PDF 파일인지 확인
                file_extension = os.path.splitext(struct_pdf_selected)[1].lower()

                if file_extension == '.xml':
                    # XML 파일 처리 (독일법)
                    from pdf_parser import extract_structured_articles_from_xml
                    st.write("독일 법령 XML 파싱 중...")
                    df_structured = extract_structured_articles_from_xml(
                        struct_pdf_selected,
                        country=struct_country,
                        law_name="특허법"  # 기본값, 필요시 UI에서 선택 가능하도록 확장 가능
                    )
                    st.write(f"{len(df_structured)}개 항목 추출 (조/항 단위)")
                elif file_extension == '.rtf':
                    # RTF 파일 처리 (미국법)
                    st.write("미국 법령 RTF 파싱 중...")
                    df_structured = extract_structured_articles(struct_pdf_selected)
                    st.write(f"{len(df_structured)}개 항목 추출 (조/항/호 단위)")
                else:
                    # PDF/RTF 파일 처리
                    df_structured = extract_structured_articles(struct_pdf_selected)
                    st.write(f"{len(df_structured)}개 항목 추출 (조/항/호 단위)")

                # 파일명 생성
                base_name = os.path.splitext(os.path.basename(struct_pdf_selected))[0]
            else:
                st.error("파일을 선택하거나 URL을 입력해주세요.")
                st.stop()

            # 구조화 파일은 구조화법률 폴더의 국가별 하위 폴더에 저장
            structured_dir = _safe_join(DATA_DIR, "output", "구조화법률")
            os.makedirs(structured_dir, exist_ok=True)

            # 파일명 생성
            if struct_country == "일본" and uploaded_file:
                # 일본 업로드 파일의 경우
                base_name_structured = f"구조화_{struct_country}_{base_name}"
            elif struct_country == "프랑스":
                # 프랑스 LEGI XML의 경우 이미 base_name이 생성됨
                base_name_structured = f"구조화_{base_name}"
            elif input_method == "HTML URL 입력" or input_method == "디렉토리 경로 입력":
                # HTML URL의 경우 이미 base_name이 생성됨
                base_name_structured = f"구조화_{base_name}"
            else:
                # 파일 확장자 제거 (.pdf 또는 .xml)
                base_name_no_ext = _basename(struct_pdf_selected).rsplit('.', 1)[0]
                base_name_structured = f"구조화_{struct_country}_{base_name_no_ext}"

            # 국가별 하위 폴더 생성 및 저장
            country = _detect_country_from_filename(base_name_structured)
            if country:
                folder_name = _country_to_folder_name(country)
                country_dir = os.path.join(structured_dir, folder_name)
                os.makedirs(country_dir, exist_ok=True)
                excel_path = os.path.join(country_dir, f"{base_name_structured}.xlsx")
            else:
                # 국가 감지 실패 시 루트 폴더에 저장
                excel_path = os.path.join(structured_dir, f"{base_name_structured}.xlsx")

            # 조문번호 정규화: 'Section N' / 'Article N' / 'Rule N' / '§ N' → 'N'
            if '조문번호' in df_structured.columns:
                df_structured['조문번호'] = df_structured['조문번호'].apply(
                    lambda x: _article_num_display(str(x)) if pd.notna(x) and str(x).strip() not in ('전문', '') else x
                )

            save_structured_to_excel(df_structured, excel_path)
            st.write(f"저장 완료: `{excel_path}`")

            status.update(label="구조화 완료", state="complete")

        # 미리보기
        st.subheader("구조화 결과 미리보기")
        st.dataframe(df_structured.head(20), use_container_width=True, hide_index=True)

        # Excel 다운로드 버튼 추가
        import io
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df_structured.to_excel(writer, index=False, sheet_name="법조문")
        excel_data = excel_buffer.getvalue()

        st.download_button(
            label="📥 구조화 Excel 다운로드",
            data=excel_data,
            file_name=os.path.basename(excel_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="structured_excel_download"
        )

# ══════════════════════════════════════════════════════════════
# 페이지 2: 번역 실행 (구조화 엑셀 → 번역 + 한국법 매칭)
# ══════════════════════════════════════════════════════════════
elif page == "번역 실행":
    st.markdown("""
    <div class="section-header">
        <h3>AI 번역 및 매칭</h3>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <p style="margin: 0; color: #64748b;">
            구조화된 법령을 Gemini 및 Claude AI로 이중 번역하고, 한국 특허법의 유사 조문을 자동으로 매칭합니다.
            번역 결과는 Excel 파일로 저장되어 상세 비교가 가능합니다.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("")  # 여백

    # ── 외국법 구조화 엑셀 선택 (나라별 필터) ──
    st.markdown("#### 외국법 선택")

    # 구조화법률 폴더의 국가별 서브폴더에서 파일 수집
    structured_dir = _safe_join(DATA_DIR, "output", "구조화법률")
    country_files: dict = {}
    if os.path.isdir(structured_dir):
        for entry in sorted(os.listdir(structured_dir)):
            if entry == "한국":
                continue
            country_dir = os.path.join(structured_dir, entry)
            if os.path.isdir(country_dir):
                files = sorted(
                    [f for f in glob.glob(os.path.join(country_dir, "*.xlsx"))
                     if not _basename(f).startswith("~$")],
                    key=os.path.getmtime, reverse=True,
                )
                if files:
                    country_files[entry] = files

    countries_with_files = list(country_files.keys())

    if not countries_with_files:
        st.warning(
            f"`{structured_dir}` 폴더에 구조화 엑셀 파일이 없습니다.\n\n"
            "'법령 구조화' 탭에서 먼저 PDF를 구조화하세요."
        )
        foreign_excel_selected = None
    else:
        col_country, col_file = st.columns([1, 2])
        with col_country:
            selected_country = st.selectbox(
                "국가 선택",
                countries_with_files,
                key="trans_country_select",
            )
        files_for_country = country_files.get(selected_country, [])
        with col_file:
            foreign_excel_selected = st.selectbox(
                "구조화 엑셀 선택",
                files_for_country,
                format_func=_basename,
                disabled=not files_for_country,
                key="trans_foreign_excel",
            )

    # 소스 언어 선택
    col_lang, col_service = st.columns(2)
    with col_lang:
        source_lang_option = st.selectbox(
            "소스 언어", ["자동 감지", "영어", "중국어"],
            key="trans_source_lang",
            help="파일명에서 자동 감지하거나, 수동으로 선택하세요.",
        )

    # 번역 서비스 선택
    with col_service:
        translation_service = st.selectbox(
            "번역 서비스",
            ["Gemini + Claude (이중 번역)", "Claude만", "Gemini만"],
            key="translation_service",
            help="사용할 번역 서비스를 선택하세요. API 키가 없는 경우 해당 서비스는 건너뜁니다.",
        )

    st.divider()

    # ── 한국법 선택 ──
    st.markdown("#### 한국법 선택 (다중 가능)")
    st.caption("구조화 엑셀과 PDF를 혼합하여 선택할 수 있습니다.")

    korea_excels = _list_korea_excels()
    korea_pdfs = _list_pdfs(KOREA_FOLDER)

    col_ke, col_kp = st.columns(2)
    with col_ke:
        korea_excel_selected = st.multiselect(
            "한국법 구조화 엑셀",
            korea_excels, format_func=_basename,
            key="trans_korea_excel",
            help="구조화된 엑셀 파일로 한국법을 로드하면 조/항/호 단위로 더 정확한 매칭이 가능합니다.",
        )
    with col_kp:
        if not korea_pdfs:
            st.warning(f"`{DATA_DIR}/{KOREA_FOLDER}/` 폴더에 한국 법령 PDF를 넣어주세요.")
        korea_pdf_selected = st.multiselect(
            "한국법 PDF",
            korea_pdfs, format_func=_basename,
            key="trans_korea_pdf",
        )

    has_korea = bool(korea_excel_selected) or bool(korea_pdf_selected)

    st.divider()

    # 테스트 모드 옵션
    test_mode = st.checkbox(
        "테스트 모드 — 처음 20조까지만 처리",
        value=False,
        key="test_mode",
        help="빠른 테스트를 위해 처음 20개 조문만 처리합니다."
    )

    btn_col1, btn_col2, btn_col3 = st.columns(3)
    with btn_col1:
        trans_run = st.button(
            "번역 실행", type="primary",
            disabled=not foreign_excel_selected or not has_korea,
            use_container_width=True, key="trans_run",
        )
    with btn_col2:
        retrans_run = st.button(
            "재번역 — 특정 조문",
            disabled=not foreign_excel_selected or not has_korea,
            use_container_width=True, key="retrans_run",
        )
    with btn_col3:
        rematch_run = st.button(
            "재매칭 — 유사 조문",
            disabled=not foreign_excel_selected or not has_korea,
            use_container_width=True, key="rematch_run",
        )

    # 번역 실행 버튼 클릭 시 session state에 저장
    if trans_run:
        st.session_state.translation_started = True
        st.session_state.retranslation_started = False
        st.session_state.rematch_started = False
        # 새로운 번역 실행 시작 - 선택 state 초기화
        if "proceed_with_choice" in st.session_state:
            del st.session_state.proceed_with_choice
        if "use_existing" in st.session_state:
            del st.session_state.use_existing
        # 이전 번역 결과 및 진행 상태 초기화
        st.session_state._trans_result = None
        st.session_state._trans_error = None
        st.session_state._trans_done = False
        st.session_state._trans_thread_active = False
        st.session_state._trans_progress = {"current": 0, "total": 0, "text": "번역 준비 중..."}
        st.session_state._trans_cancel_event = threading.Event()

    if retrans_run:
        st.session_state.retranslation_started = True
        st.session_state.translation_started = False
        st.session_state.rematch_started = False

    if rematch_run:
        st.session_state.rematch_started = True
        st.session_state.translation_started = False
        st.session_state.retranslation_started = False

    # ── 이전 번역 결과 표시 (탭 이동 후 복귀 시에도 유지) ──
    if (not st.session_state.get("translation_started")
            and not st.session_state.get("retranslation_started")
            and not st.session_state.get("rematch_started")
            and st.session_state.get("_trans_result")):
        _prev = st.session_state._trans_result
        st.success(f"번역 완료 - 저장 위치: {_prev['xlsx_path']}")
        st.dataframe(_prev["df_display"], use_container_width=True, hide_index=True)
        _c1, _c2 = st.columns(2)
        with _c1:
            st.download_button(
                "Excel 다운로드", _prev["excel_data"], f"{_prev['base_name']}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="trans_xlsx_dl_prev",
            )
        with _c2:
            if st.button("새 번역 시작 (결과 초기화)", key="clear_trans_result"):
                st.session_state._trans_result = None
                st.rerun()

    # 번역이 시작된 경우 실행
    if st.session_state.get("translation_started", False) and foreign_excel_selected:

        # ── 기존 번역 결과 확인 ──
        # 파일명에서 국가/법령명 추출
        fname = _basename(foreign_excel_selected).replace(".xlsx", "")

        # 파일명에서 국가와 법령명 추출: 구조화_국가_법령명
        parts = fname.split("_", 2)
        if len(parts) >= 3 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[2]
        elif len(parts) >= 2 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[1]
        else:
            # 비표준 형식: 파일명 전체 사용
            trans_country = fname
            trans_law_name = fname

        # 테스트 모드일 경우 파일명에 표시
        test_suffix = "_테스트" if test_mode else ""
        base_name = f"번역비교_{trans_country}_{trans_law_name}{test_suffix}"

        # 번역 스레드 실행 중/완료 시 기존 파일 UI 건너뜀
        existing_csv = None
        if not (st.session_state.get("_trans_thread_active") or st.session_state.get("_trans_done")):
            for search_dir in [_safe_join(DATA_DIR, "output"), PROJECT_DIR]:
                for ext in [".csv", ".xlsx"]:
                    candidate = _safe_join(search_dir, f"{base_name}{ext}")
                    if os.path.exists(candidate):
                        existing_csv = candidate
                        break
                if existing_csv:
                    break

        if existing_csv:
            st.warning("기존 번역 결과가 존재합니다.")

            df_existing = None
            if existing_csv.endswith((".xlsx", ".xls")):
                try:
                    df_existing = pd.read_excel(existing_csv)
                except Exception:
                    pass
            else:
                for enc in ["utf-8-sig", "utf-8"]:
                    try:
                        df_existing = pd.read_csv(existing_csv, encoding=enc)
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue

            if df_existing is not None:
                st.info(f"파일: `{_basename(existing_csv)}`")

                info_col1, info_col2, info_col3 = st.columns(3)
                with info_col1:
                    st.metric("총 조문 수", len(df_existing))
                with info_col2:
                    if "매칭 점수" in df_existing.columns:
                        scores = pd.to_numeric(df_existing["매칭 점수"], errors="coerce")
                        avg_score = scores.mean()
                        st.metric("평균 매칭 점수", f"{avg_score:.3f}" if pd.notna(avg_score) else "-")
                with info_col3:
                    if "매칭 이유" in df_existing.columns:
                        reason_count = df_existing["매칭 이유"].notna().sum()
                        st.metric("매칭 이유 제공", f"{reason_count}건")

                st.subheader("기존 결과 미리보기 (상위 5개 조문)")
                preview_df = df_existing.head(5).copy()
                for col in ["원문", "Gemini 번역", "Claude 번역"]:
                    if col in preview_df.columns:
                        preview_df[col] = preview_df[col].apply(
                            lambda x: str(x)[:100] + "..." if pd.notna(x) and len(str(x)) > 100 else x
                        )
                st.dataframe(preview_df, use_container_width=True, hide_index=True)

                st.divider()

                st.subheader("번역 실행 방식 선택")
                choice = st.radio(
                    "어떻게 진행하시겠습니까?",
                    [
                        "기존 번역 결과 사용 (빠름, 비용 절감)",
                        "새로 번역 실행 (API 호출, 시간 소요)"
                    ],
                    key="use_existing_choice",
                    help="기존 결과를 사용하면 API 호출 없이 즉시 결과를 확인할 수 있습니다."
                )
                use_existing = choice.startswith("기존")

                st.divider()
                col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
                with col_btn1:
                    if st.button("선택 확정", type="primary", use_container_width=True, key="confirm_choice_btn"):
                        st.session_state.proceed_with_choice = True
                        st.session_state.use_existing = use_existing
                with col_btn2:
                    if st.button("취소", use_container_width=True, key="cancel_choice_btn"):
                        if "proceed_with_choice" in st.session_state:
                            del st.session_state.proceed_with_choice
                        if "use_existing" in st.session_state:
                            del st.session_state.use_existing
                        st.info("분석이 취소되었습니다.")
                        st.stop()

                if not st.session_state.get("proceed_with_choice", False):
                    st.info("위에서 옵션을 선택하고 '선택 확정' 버튼을 눌러주세요.")
                    st.stop()

                # session_state에서 선택 값 가져오기
                use_existing = st.session_state.get("use_existing", True)

                if use_existing:
                    st.success(f"기존 결과 로드 완료: {len(df_existing)}건")
                    if "조문" in df_existing.columns:
                        df_existing = df_existing[df_existing["조문"] != "전문"]
                    st.dataframe(df_existing, use_container_width=True, hide_index=True, height=400)
                    st.info("'상세보기' 탭에서 상세 내용을 확인하세요.")
                    st.stop()

        # 기존 결과를 사용하지 않는 경우 새로 번역 실행
        if not st.session_state.get("use_existing", False):
            # 파일 로드 + 스레드 시작은 처음 한 번만 (폴링 rerun 시 건너뜀)
            if not st.session_state.get("_trans_thread_active") and not st.session_state.get("_trans_done"):
                # ── 1) 외국법 엑셀 로드 ──
                with st.status("외국법 구조화 엑셀 로드 중...", expanded=True) as status:
                    try:
                        df_foreign = pd.read_excel(foreign_excel_selected)
                        st.write(f"{_basename(foreign_excel_selected)} 로드: {len(df_foreign)}건")
                    except Exception as e:
                        st.error(f"엑셀 읽기 실패: {e}")
                        st.stop()

                    required_cols = ["조문번호", "원문"]
                    missing_cols = [c for c in required_cols if c not in df_foreign.columns]
                    if missing_cols:
                        st.error(f"필수 컬럼이 없습니다: {missing_cols}")
                        st.stop()

                    for col in ["편", "장", "절", "조문제목", "항", "호", "목", "세목"]:
                        if col in df_foreign.columns:
                            df_foreign[col] = df_foreign[col].fillna("").astype(str)
                        else:
                            df_foreign[col] = ""

                    foreign_articles = []
                    for _, row in df_foreign.iterrows():
                        article_id = f"{row['조문번호']}"
                        if row['항']:
                            article_id += f"-{row['항']}"
                        if row['호']:
                            article_id += f"-{row['호']}"
                        text = str(row["원문"]) if pd.notna(row["원문"]) else ""
                        if not text.strip():
                            continue
                        foreign_articles.append({
                            "id": article_id, "text": text,
                            "편": str(row.get("편", "")) if pd.notna(row.get("편")) else "",
                            "장": str(row.get("장", "")) if pd.notna(row.get("장")) else "",
                            "절": str(row.get("절", "")) if pd.notna(row.get("절")) else "",
                            "조문번호": str(row["조문번호"]) if pd.notna(row["조문번호"]) else "",
                            "조문제목": str(row.get("조문제목", "")) if pd.notna(row.get("조문제목")) else "",
                            "항": str(row.get("항", "")) if pd.notna(row.get("항")) else "",
                            "호": str(row.get("호", "")) if pd.notna(row.get("호")) else "",
                            "목": str(row.get("목", "")) if pd.notna(row.get("목")) else "",
                            "세목": str(row.get("세목", "")) if pd.notna(row.get("세목")) else "",
                        })

                    if test_mode:
                        unique_articles = []
                        seen_article_nums = set()
                        for art in foreign_articles:
                            art_num = art.get("조문번호", "")
                            if art_num not in seen_article_nums:
                                seen_article_nums.add(art_num)
                            if len(seen_article_nums) <= 20:
                                unique_articles.append(art)
                            else:
                                break
                        foreign_articles = unique_articles
                        st.info(f"테스트 모드: 처음 20개 조문만 처리합니다 (총 {len(foreign_articles)}개 항목)")

                    if source_lang_option == "영어":
                        source_lang = "english"
                    elif source_lang_option == "중국어":
                        source_lang = "chinese"
                    else:
                        source_lang = _detect_lang(foreign_excel_selected)
                    st.write(f"소스 언어: {source_lang}")
                    status.update(label="외국법 로드 완료", state="complete")

                # ── 2) 한국법 로드 (AI 매칭용) ──
                with st.status("한국 법령 로드 중...", expanded=True) as status:
                    all_korea_articles = []
                    for excel_path in korea_excel_selected:
                        try:
                            df_korea = pd.read_excel(excel_path)
                            source_name = _basename(excel_path)
                            korea_by_article = {}
                            for _, row in df_korea.iterrows():
                                article_num = row.get('조문번호', '')
                                if pd.notna(article_num) and str(article_num).strip():
                                    article_num = str(article_num)
                                    if article_num not in korea_by_article:
                                        korea_by_article[article_num] = {
                                            'rows': [],
                                            'title': str(row.get('조문제목', '')).strip() if pd.notna(row.get('조문제목')) else ""
                                        }
                                    text = str(row.get("원문", "")).strip()
                                    if text:
                                        korea_by_article[article_num]['rows'].append(text)
                            for article_num, data in korea_by_article.items():
                                combined_text = "\n".join(data['rows'])
                                all_korea_articles.append({
                                    "id": article_num, "text": combined_text,
                                    "source": source_name, "title": data['title'],
                                })
                            st.write(f"{source_name}: {len(korea_by_article)}개 조문")
                        except Exception as e:
                            st.warning(f"엑셀 읽기 실패 ({_basename(excel_path)}): {e}")

                    for kp in korea_pdf_selected:
                        k_text = parse_pdf(kp)
                        k_articles = split_articles(k_text, lang="korean")
                        for a in k_articles:
                            a["source"] = _basename(kp)
                        all_korea_articles.extend(k_articles)
                        st.write(f"{_basename(kp)}: {len(k_articles)}개 조문")

                    if not all_korea_articles:
                        st.error("한국법 조문이 없습니다. 엑셀 또는 PDF를 선택해주세요.")
                        st.stop()

                    korea_index = {"articles": all_korea_articles}
                    st.write(f"한국법 총 {len(all_korea_articles)}개 조문 로드 완료")
                    status.update(label="한국 법령 로드 완료", state="complete")

                # ── 3) 번역 실행 (백그라운드 스레드) ──
                use_gemini = "Gemini" in translation_service
                use_claude = "Claude" in translation_service

                _cancel_ev = st.session_state.get("_trans_cancel_event", threading.Event())

                # 스레드와 공유할 일반 Python dict (st.session_state 직접 접근 대신 사용)
                _shared = {
                    "progress": {"current": 0, "total": 0, "text": "번역 준비 중..."},
                    "result": None,
                    "error": None,
                    "done": False,
                }
                st.session_state._trans_shared = _shared

                # 스레드에 넘길 입력 데이터 (일반 변수로 클로저 캡처)
                _inp = {
                    "foreign_articles": foreign_articles,
                    "all_korea_articles": all_korea_articles,
                    "korea_index": korea_index,
                    "source_lang": source_lang,
                    "base_name": base_name,
                    "use_gemini": use_gemini,
                    "use_claude": use_claude,
                    "trans_country": trans_country,
                    "foreign_excel_selected": foreign_excel_selected,
                }

                def _translation_worker(_inp=_inp, _shared=_shared, _cev=_cancel_ev):
                    """백그라운드 번역 스레드 - st.session_state 직접 접근 금지"""
                    try:
                        # 진행률 콜백 (취소 체크 포함)
                        def _prog(cur, tot):
                            _shared["progress"] = {
                                "current": cur, "total": tot,
                                "text": f"번역 중... ({cur}/{tot})"
                            }
                            if _cev and _cev.is_set():
                                raise StopIteration("취소됨")

                        _translated = translate_batch(
                            _inp["foreign_articles"],
                            source_lang=_inp["source_lang"],
                            progress_callback=_prog,
                            group_by_article=True,
                            use_gemini=_inp["use_gemini"],
                            use_claude=_inp["use_claude"],
                            cancel_event=_cev,
                        )

                        if _cev and _cev.is_set():
                            _shared["error"] = "사용자가 번역을 취소했습니다."
                            return

                        _shared["progress"] = {"current": 0, "total": 0, "text": "한국법 매칭 중..."}

                        # 관련 한국법 선택
                        _korea_law_sources = sorted(set(
                            a.get("source", "") for a in _inp["all_korea_articles"] if a.get("source")
                        ))
                        _sample_text = ""
                        for item in _translated:
                            if item["id"] != "전문" and not item["id"].endswith("(삭제)"):
                                _sample_text = item.get("gemini", "") or item.get("claude", "")
                                break
                        _relevant_sources = select_relevant_korean_laws(
                            _basename(_inp["foreign_excel_selected"]), _sample_text, _korea_law_sources,
                        )

                        if _cev and _cev.is_set():
                            _shared["error"] = "사용자가 번역을 취소했습니다."
                            return

                        # 조문 단위 그룹화
                        from collections import defaultdict
                        _article_groups = defaultdict(list)
                        for item in _translated:
                            _art_num = item.get("조문번호", item["id"])
                            _article_groups[_art_num].append(item)

                        # 일괄 매칭 준비
                        _batch_articles = []
                        for _art_num, _group in _article_groups.items():
                            if _art_num != "전문" and not _art_num.endswith("(삭제)"):
                                _fi = _group[0]
                                _batch_articles.append({
                                    'id': _art_num,
                                    'text': str(_fi.get("original", "")),
                                    '조문제목': _fi.get("조문제목", ""),
                                    'translated': str(_fi.get("gemini", "")) or str(_fi.get("claude", "")),
                                })

                        _shared["progress"] = {"current": 0, "total": 0, "text": "한국법 AI 매칭 중..."}
                        _batch_results = find_similar_korean_batch(
                            _batch_articles, _inp["korea_index"],
                            relevant_law_sources=_relevant_sources,
                        )

                        # 매칭 결과 할당
                        for _art_num, _group in _article_groups.items():
                            if _art_num == "전문" or _art_num.endswith("(삭제)"):
                                for item in _group:
                                    item["similar_korean"] = []
                            else:
                                _skey = str(_art_num)
                                for _pfx in ["Article ", "Section ", "Rule ", "第"]:
                                    if _skey.startswith(_pfx):
                                        _skey = _skey[len(_pfx):]
                                        break
                                _skey = _skey.rstrip("條")
                                _matches = _batch_results.get(_skey, []) or _batch_results.get(str(_art_num), [])
                                for item in _group:
                                    item["similar_korean"] = _matches

                        # 결과 DataFrame 구성
                        _rows_full = []
                        _rows_display = []
                        for _art_num, _group in _article_groups.items():
                            _fi = _group[0]
                            _best_match = _best_score = _best_kt = _best_reason = ""
                            if _fi.get("similar_korean"):
                                _top = _fi["similar_korean"][0]
                                _lname = _korean_law_name(_top.get("source", ""))
                                _kid = _top['korean_id']
                                if not _kid.startswith("제"):
                                    _kid = f"제{_kid}조"
                                _best_match = f"{_lname} {_kid}"
                                _best_score = f"{_top['score']:.3f}"
                                _best_kt = _top.get("korean_text", "")
                                _best_reason = _top.get("ai_reason", "")

                            _combined_orig = str(_fi.get("original", ""))
                            _combined_gem = str(_fi.get("gemini", ""))
                            _combined_cla = str(_fi.get("claude", ""))
                            _row = {"국가": _inp["trans_country"]}
                            for _k in ["편", "장", "절", "조문번호", "조문제목"]:
                                _row[_k] = _fi.get(_k, "")

                            _rows_full.append({
                                **_row, "원문": _combined_orig,
                                "Gemini 번역": _combined_gem, "Claude 번역": _combined_cla,
                                "유사 한국법": _best_match, "매칭 점수": _best_score,
                                "한국법 조문 내용": _best_kt, "매칭 이유": _best_reason,
                            })
                            _rows_display.append({
                                **_row,
                                "원문": _combined_orig[:200] + ("..." if len(_combined_orig) > 200 else ""),
                                "Gemini 번역": _combined_gem[:200] + ("..." if len(_combined_gem) > 200 else ""),
                                "Claude 번역": _combined_cla[:200] + ("..." if len(_combined_cla) > 200 else ""),
                                "유사 한국법": _best_match, "매칭 이유": _best_reason,
                            })

                        _df_full = pd.DataFrame(_rows_full)
                        _df_display = pd.DataFrame(_rows_display)

                        # Excel 저장
                        _trans_dir = _safe_join(DATA_DIR, "output", "번역비교결과")
                        os.makedirs(_trans_dir, exist_ok=True)
                        _fparts = _inp["foreign_excel_selected"].replace("\\", "/").split("/")
                        _cfolder = None
                        if "구조화법률" in _fparts:
                            _ci = _fparts.index("구조화법률")
                            if _ci + 1 < len(_fparts):
                                _cfolder = _fparts[_ci + 1]
                        if _cfolder and _cfolder in COUNTRY_MAP:
                            _cdir = os.path.join(_trans_dir, _cfolder)
                            os.makedirs(_cdir, exist_ok=True)
                            _xlsx_path = os.path.join(_cdir, f"{_inp['base_name']}.xlsx")
                        else:
                            _xlsx_path = os.path.join(_trans_dir, f"{_inp['base_name']}.xlsx")

                        _buf = io.BytesIO()
                        with pd.ExcelWriter(_buf, engine="openpyxl") as _ew:
                            _df_full.to_excel(_ew, index=False, sheet_name="번역결과")
                            _ws = _ew.sheets["번역결과"]
                            for _row_cells in _ws.iter_rows():
                                for _cell in _row_cells:
                                    _cell.alignment = Alignment(wrap_text=True, vertical='top')
                        _excel_data = _buf.getvalue()
                        with open(_xlsx_path, "wb") as _f:
                            _f.write(_excel_data)

                        _shared["result"] = {
                            "df_display": _df_display,
                            "excel_data": _excel_data,
                            "base_name": _inp["base_name"],
                            "xlsx_path": _xlsx_path,
                        }

                    except StopIteration:
                        _shared["error"] = "사용자가 번역을 취소했습니다."
                    except Exception as _e:
                        _shared["error"] = str(_e)
                    finally:
                        _shared["done"] = True

                _t = threading.Thread(target=_translation_worker, daemon=True)
                _t.start()
                st.session_state._trans_thread_active = True
                # 파일 로드 화면을 지우고 깨끗한 진행 화면으로 전환
                st.rerun()

            # ── 진행 중: 프로그레스바 + 취소 버튼 표시 ──
            _shared = st.session_state.get("_trans_shared", {})
            if not _shared.get("done"):
                _prog = _shared.get("progress", {})
                _cur = _prog.get("current", 0)
                _tot = max(_prog.get("total", 1), 1)
                st.subheader("번역 진행")
                st.progress(_cur / _tot, text=_prog.get("text", "번역 중..."))

                if st.button("번역 취소", key="cancel_trans_btn", type="secondary"):
                    _cev2 = st.session_state.get("_trans_cancel_event")
                    if _cev2:
                        _cev2.set()
                    st.warning("취소 요청이 전송되었습니다. 현재 조문 처리 완료 후 중단됩니다.")
                    time.sleep(0.5)
                    st.rerun()

                # 완료될 때까지 폴링
                time.sleep(0.5)
                st.rerun()
                st.stop()

            # ── 완료: shared → session_state로 복사 후 rerun ──
            st.session_state.translation_started = False
            st.session_state._trans_done = True
            st.session_state._trans_thread_active = False
            if _shared.get("result"):
                st.session_state._trans_result = _shared["result"]
                st.rerun()  # 이전 결과 표시 블록(상단)으로 전환
            if _shared.get("error"):
                st.error(_shared["error"])
            st.stop()

    # ══════════════════════════════════════════════════════════════
    # 재번역 모드
    # ══════════════════════════════════════════════════════════════
    elif st.session_state.get("retranslation_started", False) and foreign_excel_selected:

        # ── 기존 번역결과 파일 찾기 ──
        fname = _basename(foreign_excel_selected).replace(".xlsx", "")

        # 파일명에서 국가와 법령명 추출: 구조화_국가_법령명
        parts = fname.split("_", 2)
        if len(parts) >= 3 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[2]
        elif len(parts) >= 2 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[1]
        else:
            trans_country = fname
            trans_law_name = fname

        # 기존 번역결과 파일 검색 (테스트 포함)
        existing_result = None
        for test_suffix in ["", "_테스트"]:
            base_name = f"번역비교_{trans_country}_{trans_law_name}{test_suffix}"
            for search_dir in [
                _safe_join(DATA_DIR, "output", "번역비교결과"),
                _safe_join(DATA_DIR, "output"),
                PROJECT_DIR,
            ]:
                for ext in [".xlsx", ".csv"]:
                    candidate = _safe_join(search_dir, f"{base_name}{ext}")
                    if os.path.exists(candidate):
                        existing_result = candidate
                        break
                if existing_result:
                    break
            if existing_result:
                break

        if not existing_result:
            st.error("기존 번역결과 파일을 찾을 수 없습니다. 먼저 '번역 실행'을 해주세요.")
            st.stop()

        # ── 구조화 엑셀 및 기존 번역결과 로드 ──
        try:
            df_foreign = pd.read_excel(foreign_excel_selected)
        except Exception as e:
            st.error(f"구조화 엑셀 읽기 실패: {e}")
            st.stop()

        try:
            if existing_result.endswith((".xlsx", ".xls")):
                df_existing = pd.read_excel(existing_result)
            else:
                df_existing = pd.read_csv(existing_result, encoding="utf-8-sig")
        except Exception as e:
            st.error(f"기존 번역결과 읽기 실패: {e}")
            st.stop()

        st.info(f"구조화 엑셀: `{_basename(foreign_excel_selected)}`\n\n"
                f"기존 번역결과: `{_basename(existing_result)}`")

        # ── 조문 목록 표시 (체크박스) ──
        st.subheader("재번역할 조문 선택")
        st.caption("구조화 엑셀에서 수정한 조문을 체크하세요. 선택한 조문만 재번역 + 재매칭됩니다.")

        # 조문번호별로 그룹화 (중복 조문번호 하나로 표시)
        article_nums_seen = []
        article_info = {}
        for _, row in df_foreign.iterrows():
            art_num = str(row.get("조문번호", ""))
            if not art_num or art_num in article_info:
                continue
            art_title = str(row.get("조문제목", "")) if pd.notna(row.get("조문제목")) else ""
            article_nums_seen.append(art_num)
            article_info[art_num] = art_title

        # 전체 선택 / 해제
        def _retrans_toggle_all():
            val = st.session_state.retrans_select_all
            for a_num in article_info:
                st.session_state[f"retrans_chk_{a_num}"] = val

        st.checkbox("전체 선택", value=False, key="retrans_select_all", on_change=_retrans_toggle_all)

        selected_articles = []
        cols_per_row = 3
        article_list = list(article_info.items())

        for i in range(0, len(article_list), cols_per_row):
            cols = st.columns(cols_per_row)
            for j, col in enumerate(cols):
                idx = i + j
                if idx >= len(article_list):
                    break
                art_num, art_title = article_list[idx]
                label = _article_num_display(art_num)
                if art_title:
                    label += f" ({art_title})"
                with col:
                    if st.checkbox(label, key=f"retrans_chk_{art_num}"):
                        selected_articles.append(art_num)

        if not selected_articles:
            st.warning("재번역할 조문을 선택하세요.")
            st.stop()

        st.success(f"선택된 조문: {len(selected_articles)}개")

        # ── 재번역 실행 버튼 ──
        retrans_execute = st.button(
            f"선택한 {len(selected_articles)}개 조문 재번역 + 재매칭 실행",
            type="primary", use_container_width=True, key="retrans_execute",
        )

        if retrans_execute:
            # ── 1) 선택한 조문만 구조화 엑셀에서 추출 ──
            with st.status("선택한 조문 로드 중...", expanded=True) as status:
                # NaN 채우기
                for col in ["편", "장", "절", "조문제목", "항", "호", "목", "세목"]:
                    if col in df_foreign.columns:
                        df_foreign[col] = df_foreign[col].fillna("").astype(str)
                    else:
                        df_foreign[col] = ""

                retrans_articles = []
                for _, row in df_foreign.iterrows():
                    art_num = str(row.get("조문번호", ""))
                    if art_num not in selected_articles:
                        continue

                    article_id = f"{row['조문번호']}"
                    if row['항']:
                        article_id += f"-{row['항']}"
                    if row['호']:
                        article_id += f"-{row['호']}"

                    text = str(row["원문"]) if pd.notna(row["원문"]) else ""
                    if not text.strip():
                        continue

                    retrans_articles.append({
                        "id": article_id,
                        "text": text,
                        "편": str(row.get("편", "")) if pd.notna(row.get("편")) else "",
                        "장": str(row.get("장", "")) if pd.notna(row.get("장")) else "",
                        "절": str(row.get("절", "")) if pd.notna(row.get("절")) else "",
                        "조문번호": str(row["조문번호"]) if pd.notna(row["조문번호"]) else "",
                        "조문제목": str(row.get("조문제목", "")) if pd.notna(row.get("조문제목")) else "",
                        "항": str(row.get("항", "")) if pd.notna(row.get("항")) else "",
                        "호": str(row.get("호", "")) if pd.notna(row.get("호")) else "",
                        "목": str(row.get("목", "")) if pd.notna(row.get("목")) else "",
                        "세목": str(row.get("세목", "")) if pd.notna(row.get("세목")) else "",
                    })

                # 소스 언어 결정
                if source_lang_option == "영어":
                    source_lang = "english"
                elif source_lang_option == "중국어":
                    source_lang = "chinese"
                else:
                    source_lang = _detect_lang(foreign_excel_selected)

                st.write(f"{len(retrans_articles)}개 항목 로드 (소스 언어: {source_lang})")
                status.update(label="조문 로드 완료", state="complete")

            # ── 2) 한국법 로드 ──
            with st.status("한국 법령 로드 중...", expanded=True) as status:
                all_korea_articles = []

                for excel_path in korea_excel_selected:
                    try:
                        df_korea = pd.read_excel(excel_path)
                        source_name = _basename(excel_path)

                        korea_by_article = {}
                        for _, row in df_korea.iterrows():
                            article_num = row.get('조문번호', '')
                            if pd.notna(article_num) and str(article_num).strip():
                                article_num = str(article_num)
                                if article_num not in korea_by_article:
                                    korea_by_article[article_num] = {
                                        'rows': [],
                                        'title': str(row.get('조문제목', '')).strip() if pd.notna(row.get('조문제목')) else ""
                                    }
                                text = str(row.get("원문", "")).strip()
                                if text:
                                    korea_by_article[article_num]['rows'].append(text)

                        for article_num, data in korea_by_article.items():
                            combined_text = "\n".join(data['rows'])
                            all_korea_articles.append({
                                "id": article_num,
                                "text": combined_text,
                                "source": source_name,
                                "title": data['title'],
                            })

                        st.write(f"{source_name}: {len(korea_by_article)}개 조문")
                    except Exception as e:
                        st.warning(f"엑셀 읽기 실패 ({_basename(excel_path)}): {e}")

                for kp in korea_pdf_selected:
                    k_text = parse_pdf(kp)
                    k_articles = split_articles(k_text, lang="korean")
                    for a in k_articles:
                        a["source"] = _basename(kp)
                    all_korea_articles.extend(k_articles)
                    st.write(f"{_basename(kp)}: {len(k_articles)}개 조문")

                if not all_korea_articles:
                    st.error("한국법 조문이 없습니다.")
                    st.stop()

                korea_index = {"articles": all_korea_articles}
                st.write(f"한국법 총 {len(all_korea_articles)}개 조문 로드 완료")
                status.update(label="한국 법령 로드 완료", state="complete")

            # ── 3) 선택한 조문만 재번역 ──
            st.subheader("재번역 진행")
            progress_bar = st.progress(0, text="재번역 준비 중...")

            def _update_progress(current, total):
                progress_bar.progress(current / total, text=f"재번역 중... ({current}/{total})")

            use_gemini = "Gemini" in translation_service
            use_claude = "Claude" in translation_service

            translated = translate_batch(
                retrans_articles,
                source_lang=source_lang,
                progress_callback=_update_progress,
                group_by_article=True,
                use_gemini=use_gemini,
                use_claude=use_claude,
            )
            progress_bar.progress(1.0, text="재번역 완료!")

            # ── 4) 선택한 조문만 재매칭 ──
            st.subheader("한국법 재매칭")

            with st.status("관련 한국법 선택 중...", expanded=True) as status:
                korea_law_sources = sorted(set(
                    a.get("source", "") for a in all_korea_articles if a.get("source")
                ))
                sample_text = ""
                for item in translated:
                    if item["id"] != "전문" and not item["id"].endswith("(삭제)"):
                        sample_text = item.get("gemini", "") or item.get("claude", "")
                        break

                relevant_sources = select_relevant_korean_laws(
                    _basename(foreign_excel_selected), sample_text, korea_law_sources,
                )
                for src in relevant_sources:
                    st.write(f"관련 한국법: {_korean_law_name(src)}")
                status.update(label=f"관련 한국법 {len(relevant_sources)}개 선택 완료", state="complete")

            from collections import defaultdict
            article_groups = defaultdict(list)
            for item in translated:
                article_num = item.get("조문번호", item["id"])
                article_groups[article_num].append(item)

            match_progress = st.progress(0, text="한국법 조문 재매칭 중...")

            batch_articles = []
            for article_num, group in article_groups.items():
                if article_num != "전문" and not article_num.endswith("(삭제)"):
                    first_item = group[0]
                    batch_articles.append({
                        'id': article_num,
                        'text': str(first_item.get("original", "")),
                        '조문제목': first_item.get("조문제목", ""),
                        'translated': str(first_item.get("gemini", "")) or str(first_item.get("claude", ""))
                    })

            match_progress.progress(0.5, text="한국법 조문 재매칭 중... (AI 처리 중)")
            batch_results = find_similar_korean_batch(
                batch_articles, korea_index, relevant_law_sources=relevant_sources
            )

            for article_num, group in article_groups.items():
                if article_num == "전문" or article_num.endswith("(삭제)"):
                    for item in group:
                        item["similar_korean"] = []
                else:
                    search_key = str(article_num)
                    for prefix in ["Article ", "Section ", "Rule ", "第"]:
                        if search_key.startswith(prefix):
                            search_key = search_key[len(prefix):]
                            break
                    search_key = search_key.rstrip("條")
                    matches = batch_results.get(search_key, [])
                    if not matches:
                        matches = batch_results.get(str(article_num), [])
                    for item in group:
                        item["similar_korean"] = matches

            match_progress.progress(1.0, text="재매칭 완료!")

            # ── 5) 기존 번역결과 Excel 업데이트 ──
            st.subheader("결과 업데이트")

            # 재번역된 조문으로 새 행 생성
            new_rows = {}
            for article_num, group in article_groups.items():
                first_item = group[0]

                best_match = ""
                best_score = ""
                best_korean_text = ""
                best_reason = ""
                if first_item.get("similar_korean"):
                    top = first_item["similar_korean"][0]
                    law_name = _korean_law_name(top.get("source", ""))
                    korean_id = top['korean_id']
                    if not korean_id.startswith("제"):
                        korean_id = f"제{korean_id}조"
                    best_match = f"{law_name} {korean_id}"
                    best_score = f"{top['score']:.3f}"
                    best_korean_text = top.get("korean_text", "")
                    best_reason = top.get("ai_reason", "")

                combined_original = str(first_item.get("original", ""))
                combined_gemini = str(first_item.get("gemini", ""))
                combined_claude = str(first_item.get("claude", ""))

                new_rows[str(article_num)] = {
                    "국가": trans_country,
                    "편": first_item.get("편", ""),
                    "장": first_item.get("장", ""),
                    "절": first_item.get("절", ""),
                    "조문번호": first_item.get("조문번호", ""),
                    "조문제목": first_item.get("조문제목", ""),
                    "원문": combined_original,
                    "Gemini 번역": combined_gemini,
                    "Claude 번역": combined_claude,
                    "유사 한국법": best_match,
                    "매칭 점수": best_score,
                    "한국법 조문 내용": best_korean_text,
                    "매칭 이유": best_reason,
                }

            # 기존 DataFrame에서 해당 조문 행 교체
            updated_count = 0
            for idx, row in df_existing.iterrows():
                art_num = str(row.get("조문번호", ""))
                if art_num in new_rows:
                    for col, val in new_rows[art_num].items():
                        if col in df_existing.columns:
                            df_existing.at[idx, col] = val
                    updated_count += 1

            # 기존 파일에 덮어쓰기 저장
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                df_existing.to_excel(writer, index=False, sheet_name="번역결과")

                # 셀 포맷팅: 줄바꿈 표시
                worksheet = writer.sheets["번역결과"]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            excel_data = excel_buffer.getvalue()

            with open(existing_result, "wb") as f:
                f.write(excel_data)

            st.success(
                f"재번역 완료 — {updated_count}개 조문이 업데이트되었습니다.\n\n"
                f"저장: `{_basename(existing_result)}`"
            )

            # 업데이트된 조문 미리보기
            st.subheader("업데이트된 조문")
            updated_df = df_existing[df_existing["조문번호"].astype(str).isin(selected_articles)]
            if not updated_df.empty:
                display_cols = ["조문번호", "조문제목", "원문", "Gemini 번역", "Claude 번역", "유사 한국법"]
                display_cols = [c for c in display_cols if c in updated_df.columns]
                preview = updated_df[display_cols].copy()
                for col in ["원문", "Gemini 번역", "Claude 번역"]:
                    if col in preview.columns:
                        preview[col] = preview[col].apply(
                            lambda x: str(x)[:150] + "..." if pd.notna(x) and len(str(x)) > 150 else x
                        )
                st.dataframe(preview, use_container_width=True, hide_index=True)

            st.download_button(
                "업데이트된 Excel 다운로드", excel_data,
                _basename(existing_result),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="retrans_dl",
            )

    # ══════════════════════════════════════════════════════════════
    # 재매칭 모드 (번역은 유지, 유사 조문 매칭만 다시)
    # ══════════════════════════════════════════════════════════════
    elif st.session_state.get("rematch_started", False) and foreign_excel_selected:

        # ── 기존 번역결과 파일 찾기 ──
        fname = _basename(foreign_excel_selected).replace(".xlsx", "")

        # 파일명에서 국가와 법령명 추출: 구조화_국가_법령명
        parts = fname.split("_", 2)
        if len(parts) >= 3 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[2]
        elif len(parts) >= 2 and parts[0] == "구조화":
            trans_country = parts[1]
            trans_law_name = parts[1]
        else:
            trans_country = fname
            trans_law_name = fname

        existing_result = None
        for test_suffix in ["", "_테스트"]:
            base_name = f"번역비교_{trans_country}_{trans_law_name}{test_suffix}"
            for search_dir in [
                _safe_join(DATA_DIR, "output", "번역비교결과"),
                _safe_join(DATA_DIR, "output"),
                PROJECT_DIR,
            ]:
                for ext in [".xlsx", ".csv"]:
                    candidate = _safe_join(search_dir, f"{base_name}{ext}")
                    if os.path.exists(candidate):
                        existing_result = candidate
                        break
                if existing_result:
                    break
            if existing_result:
                break

        if not existing_result:
            st.error("기존 번역결과 파일을 찾을 수 없습니다. 먼저 '번역 실행'을 해주세요.")
            st.stop()

        # ── 기존 번역결과 로드 ──
        try:
            if existing_result.endswith((".xlsx", ".xls")):
                df_existing = pd.read_excel(existing_result)
            else:
                df_existing = pd.read_csv(existing_result, encoding="utf-8-sig")
        except Exception as e:
            st.error(f"기존 번역결과 읽기 실패: {e}")
            st.stop()

        st.info(f"기존 번역결과: `{_basename(existing_result)}`")

        # ── 조문 목록 표시 (체크박스) ──
        st.subheader("재매칭할 조문 선택")
        st.caption("유사 한국법 매칭을 다시 실행할 조문을 체크하세요. 번역은 그대로 유지됩니다.")

        # 조문번호 목록 추출
        article_info = {}
        if "조문번호" in df_existing.columns:
            for _, row in df_existing.iterrows():
                art_num = str(row.get("조문번호", ""))
                if not art_num or art_num in article_info:
                    continue
                art_title = str(row.get("조문제목", "")) if pd.notna(row.get("조문제목")) else ""
                article_info[art_num] = art_title

        if not article_info:
            st.error("번역결과에서 조문 정보를 찾을 수 없습니다.")
            st.stop()

        def _rematch_toggle_all():
            val = st.session_state.rematch_select_all
            for a_num in article_info:
                st.session_state[f"rematch_chk_{a_num}"] = val

        st.checkbox("전체 선택", value=False, key="rematch_select_all", on_change=_rematch_toggle_all)

        selected_articles = []
        cols_per_row = 3
        article_list = list(article_info.items())

        for i in range(0, len(article_list), cols_per_row):
            cols = st.columns(cols_per_row)
            for j, col in enumerate(cols):
                idx = i + j
                if idx >= len(article_list):
                    break
                art_num, art_title = article_list[idx]
                label = _article_num_display(art_num)
                if art_title:
                    label += f" ({art_title})"
                with col:
                    if st.checkbox(label, key=f"rematch_chk_{art_num}"):
                        selected_articles.append(art_num)

        if not selected_articles:
            st.warning("재매칭할 조문을 선택하세요.")
            st.stop()

        st.success(f"선택된 조문: {len(selected_articles)}개")

        rematch_execute = st.button(
            f"선택한 {len(selected_articles)}개 조문 재매칭 실행",
            type="primary", use_container_width=True, key="rematch_execute",
        )

        if rematch_execute:
            # ── 1) 한국법 로드 ──
            with st.status("한국 법령 로드 중...", expanded=True) as status:
                all_korea_articles = []

                for excel_path in korea_excel_selected:
                    try:
                        df_korea = pd.read_excel(excel_path)
                        source_name = _basename(excel_path)

                        korea_by_article = {}
                        for _, row in df_korea.iterrows():
                            article_num = row.get('조문번호', '')
                            if pd.notna(article_num) and str(article_num).strip():
                                article_num = str(article_num)
                                if article_num not in korea_by_article:
                                    korea_by_article[article_num] = {
                                        'rows': [],
                                        'title': str(row.get('조문제목', '')).strip() if pd.notna(row.get('조문제목')) else ""
                                    }
                                text = str(row.get("원문", "")).strip()
                                if text:
                                    korea_by_article[article_num]['rows'].append(text)

                        for article_num, data in korea_by_article.items():
                            combined_text = "\n".join(data['rows'])
                            all_korea_articles.append({
                                "id": article_num,
                                "text": combined_text,
                                "source": source_name,
                                "title": data['title'],
                            })

                        st.write(f"{source_name}: {len(korea_by_article)}개 조문")
                    except Exception as e:
                        st.warning(f"엑셀 읽기 실패 ({_basename(excel_path)}): {e}")

                for kp in korea_pdf_selected:
                    k_text = parse_pdf(kp)
                    k_articles = split_articles(k_text, lang="korean")
                    for a in k_articles:
                        a["source"] = _basename(kp)
                    all_korea_articles.extend(k_articles)
                    st.write(f"{_basename(kp)}: {len(k_articles)}개 조문")

                if not all_korea_articles:
                    st.error("한국법 조문이 없습니다.")
                    st.stop()

                korea_index = {"articles": all_korea_articles}
                st.write(f"한국법 총 {len(all_korea_articles)}개 조문 로드 완료")
                status.update(label="한국 법령 로드 완료", state="complete")

            # ── 2) 선택한 조문의 번역문으로 재매칭 ──
            st.subheader("유사 조문 재매칭")

            with st.status("관련 한국법 선택 중...", expanded=True) as status:
                korea_law_sources = sorted(set(
                    a.get("source", "") for a in all_korea_articles if a.get("source")
                ))

                # 기존 번역문에서 샘플 텍스트 가져오기
                sample_text = ""
                for _, row in df_existing.iterrows():
                    art_num = str(row.get("조문번호", ""))
                    if art_num in selected_articles:
                        sample_text = str(row.get("Gemini 번역", "")) or str(row.get("Claude 번역", ""))
                        if sample_text:
                            break

                relevant_sources = select_relevant_korean_laws(
                    _basename(foreign_excel_selected), sample_text, korea_law_sources,
                )
                for src in relevant_sources:
                    st.write(f"관련 한국법: {_korean_law_name(src)}")
                status.update(label=f"관련 한국법 {len(relevant_sources)}개 선택 완료", state="complete")

            # 선택한 조문만 매칭 대상으로 구성
            batch_articles = []
            for _, row in df_existing.iterrows():
                art_num = str(row.get("조문번호", ""))
                if art_num not in selected_articles:
                    continue
                # 이미 처리한 조문번호는 스킵 (중복 방지)
                if any(b['id'] == art_num for b in batch_articles):
                    continue

                batch_articles.append({
                    'id': art_num,
                    'text': str(row.get("원문", "")) if pd.notna(row.get("원문")) else "",
                    '조문제목': str(row.get("조문제목", "")) if pd.notna(row.get("조문제목")) else "",
                    'translated': str(row.get("Gemini 번역", "")) or str(row.get("Claude 번역", ""))
                })

            match_progress = st.progress(0, text="한국법 조문 재매칭 중...")
            match_progress.progress(0.5, text="한국법 조문 재매칭 중... (AI 처리 중)")

            batch_results = find_similar_korean_batch(
                batch_articles, korea_index, relevant_law_sources=relevant_sources
            )

            match_progress.progress(1.0, text="재매칭 완료!")

            # ── 3) 기존 번역결과 Excel 업데이트 (매칭 컬럼만) ──
            st.subheader("결과 업데이트")

            updated_count = 0
            for idx, row in df_existing.iterrows():
                art_num = str(row.get("조문번호", ""))
                if art_num not in selected_articles:
                    continue

                # 매칭 결과 찾기
                search_key = art_num
                for prefix in ["Article ", "Section ", "Rule ", "第"]:
                    if search_key.startswith(prefix):
                        search_key = search_key[len(prefix):]
                        break
                search_key = search_key.rstrip("條")

                matches = batch_results.get(search_key, [])
                if not matches:
                    matches = batch_results.get(art_num, [])

                if matches:
                    top = matches[0]
                    law_name = _korean_law_name(top.get("source", ""))
                    korean_id = top['korean_id']
                    if not korean_id.startswith("제"):
                        korean_id = f"제{korean_id}조"

                    if "유사 한국법" in df_existing.columns:
                        df_existing.at[idx, "유사 한국법"] = f"{law_name} {korean_id}"
                    if "매칭 점수" in df_existing.columns:
                        df_existing.at[idx, "매칭 점수"] = f"{top['score']:.3f}"
                    if "한국법 조문 내용" in df_existing.columns:
                        df_existing.at[idx, "한국법 조문 내용"] = top.get("korean_text", "")
                    if "매칭 이유" in df_existing.columns:
                        df_existing.at[idx, "매칭 이유"] = top.get("ai_reason", "")
                else:
                    if "유사 한국법" in df_existing.columns:
                        df_existing.at[idx, "유사 한국법"] = ""
                    if "매칭 점수" in df_existing.columns:
                        df_existing.at[idx, "매칭 점수"] = ""
                    if "한국법 조문 내용" in df_existing.columns:
                        df_existing.at[idx, "한국법 조문 내용"] = ""
                    if "매칭 이유" in df_existing.columns:
                        df_existing.at[idx, "매칭 이유"] = ""

                updated_count += 1

            # 저장
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                df_existing.to_excel(writer, index=False, sheet_name="번역결과")

                # 셀 포맷팅: 줄바꿈 표시
                worksheet = writer.sheets["번역결과"]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            excel_data = excel_buffer.getvalue()

            with open(existing_result, "wb") as f:
                f.write(excel_data)

            st.success(
                f"재매칭 완료 — {updated_count}개 조문의 유사 한국법 매칭이 업데이트되었습니다.\n\n"
                f"저장: `{_basename(existing_result)}`"
            )

            # 업데이트된 조문 미리보기
            st.subheader("업데이트된 매칭 결과")
            updated_df = df_existing[df_existing["조문번호"].astype(str).isin(selected_articles)]
            if not updated_df.empty:
                display_cols = ["조문번호", "조문제목", "유사 한국법", "매칭 점수", "매칭 이유"]
                display_cols = [c for c in display_cols if c in updated_df.columns]
                st.dataframe(updated_df[display_cols], use_container_width=True, hide_index=True)

            st.download_button(
                "업데이트된 Excel 다운로드", excel_data,
                _basename(existing_result),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="rematch_dl",
            )

    elif not trans_run:
        pass

# ══════════════════════════════════════════════════════════════
# 페이지 3: 번역결과 상세보기
# ══════════════════════════════════════════════════════════════
else:  # page == "상세보기"
    st.markdown("""
    <div class="section-header">
        <h3>번역 결과 상세보기</h3>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <p style="margin: 0; color: #64748b;">
            번역된 법령 조문을 원문, Gemini 번역, Claude 번역, 한국법 매칭 정보와 함께 비교 분석합니다.
            전체보기 또는 조문별 상세보기 모드를 선택할 수 있습니다.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("")  # 여백

    result_files = _list_result_files()

    # 국가 목록: 실제 폴더 기준 (파일 유무 무관, 새 폴더 추가 시 자동 반영)
    _translation_dir = _safe_join(DATA_DIR, "output", "번역비교결과")
    _all_countries = []
    if os.path.isdir(_translation_dir):
        _all_countries = sorted([
            d for d in os.listdir(_translation_dir)
            if os.path.isdir(os.path.join(_translation_dir, d)) and d != "한국"
        ])

    # 파일을 국가별로 분류
    _detail_country_files: dict = {c: [] for c in _all_countries}
    for _f in result_files:
        _rel = _f.replace("\\", "/")
        _parts = _rel.split("/")
        if "번역비교결과" in _parts:
            _idx = _parts.index("번역비교결과")
            _c = _parts[_idx + 1] if _idx + 1 < len(_parts) - 1 else "(기타)"
        else:
            _c = "(기타)"
        if _c in _detail_country_files:
            _detail_country_files[_c].append(_f)
        else:
            _detail_country_files.setdefault(_c, []).append(_f)

    if not _all_countries:
        st.warning(
            "조회할 번역 결과 파일이 없습니다.\n\n"
            "'번역 실행' 탭에서 먼저 번역을 진행해주세요."
        )
    else:
        _col_c, _col_f = st.columns([1, 2])
        with _col_c:
            _sel_country = st.selectbox("국가 선택", _all_countries, key="detail_country_select")
        _files_for_detail = _detail_country_files.get(_sel_country, [])
        with _col_f:
            selected_file = st.selectbox(
                "결과 파일 선택",
                _files_for_detail if _files_for_detail else ["(파일 없음)"],
                format_func=_basename,
                disabled=not _files_for_detail,
                key="csv_viewer_select",
            )
        selected_file = selected_file if _files_for_detail else None

        if selected_file:
            df_csv = None

            if selected_file:
                if selected_file.endswith((".xlsx", ".xls")):
                    try:
                        df_csv = pd.read_excel(selected_file)
                    except Exception as e:
                        st.error(f"Excel 파일을 읽을 수 없습니다: {e}")
                else:
                    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
                        try:
                            df_csv = pd.read_csv(selected_file, encoding=enc)
                            break
                        except (UnicodeDecodeError, UnicodeError):
                            continue

            if df_csv is None:
                st.error("파일을 읽을 수 없습니다. 형식이나 인코딩을 확인해주세요.")
            else:
                display_name = _basename(selected_file)
                st.subheader(f"{display_name}")

                # 요약 정보
                col_info1, col_info2, col_info3 = st.columns(3)
                with col_info1:
                    st.metric("총 조문 수", len(df_csv))
                with col_info2:
                    if "국가" in df_csv.columns:
                        st.metric("국가", df_csv["국가"].iloc[0] if len(df_csv) > 0 else "-")
                with col_info3:
                    if "매칭 점수" in df_csv.columns:
                        scores = pd.to_numeric(df_csv["매칭 점수"], errors="coerce")
                        avg_score = scores.mean()
                        st.metric("평균 매칭 점수", f"{avg_score:.3f}" if pd.notna(avg_score) else "-")

                st.divider()

                df_filtered = df_csv.copy()

                # '전문' 행 제외
                if "조문" in df_filtered.columns:
                    df_filtered = df_filtered[df_filtered["조문"] != "전문"]
                elif "조문번호" in df_filtered.columns:
                    # 구조화 데이터는 전문이 따로 없으므로 스킵
                    pass

                st.caption(f"표시 중: {len(df_filtered)}건 / 전체 {len(df_csv)}건")

                # 보기 모드 선택
                view_mode = st.radio(
                    "보기 모드", ["조문별 상세 보기", "전체 보기 (복사용)"],
                    horizontal=True, key="csv_view_mode",
                )

                st.markdown(DETAIL_STYLE, unsafe_allow_html=True)

                csv_name = display_name.replace(".csv", "").replace(".xlsx", "")
                parts = csv_name.split("_", 2)
                foreign_law_name = parts[2] if len(parts) >= 3 else csv_name

                if view_mode == "전체 보기 (복사용)":
                    # ── 전체 보기: 3열 정렬 테이블 ──
                    st.subheader("전체 보기")

                    # 매칭 정보 컬럼 존재 여부 확인
                    has_matching = "유사 한국법" in df_filtered.columns

                    table_html = """<style>
.fullview-table { width:100%; border-collapse:collapse; table-layout:fixed; font-family:inherit; }
.fullview-table th { background:#f0f0f0; padding:10px 12px; border:1px solid #ccc;
    text-align:left; position:sticky; top:0; z-index:1; font-size:0.9em; }
.fullview-table td { padding:10px 12px; border:1px solid #ccc; vertical-align:top;
    white-space:pre-wrap; word-break:break-word; font-size:0.88em; line-height:1.6; }
.fullview-table tr:nth-child(even) td { background:#f9f9f9; }
.col-id { width:8%; } .col-text { width:30%; } .col-text-narrow { width:22%; } .col-korean { width:22%; }
</style>
<table class="fullview-table">
                    <colgroup>
                        <col class="col-id">"""

                    if has_matching:
                        # 한국법이 있을 때: 각 텍스트 컬럼 너비를 줄임
                        table_html += '<col class="col-text-narrow"><col class="col-text-narrow"><col class="col-text-narrow"><col class="col-korean">'
                    else:
                        # 한국법이 없을 때: 기존 너비 유지
                        table_html += '<col class="col-text"><col class="col-text"><col class="col-text">'

                    table_html += """
                    </colgroup>
                    <thead><tr>
                        <th>조문</th><th style="color:#8b2240">원문</th>
                        <th style="color:#b8860b">Gemini 번역</th>
                        <th style="color:#6e1a33">Claude 번역</th>"""

                    if has_matching:
                        table_html += '<th style="color:#a0522d">유사 한국법</th>'

                    table_html += """
                    </tr></thead><tbody>"""

                    full_original = []
                    full_gemini = []
                    full_claude = []
                    full_korean = []

                    for _, row in df_filtered.iterrows():
                        # 조문 ID 구성: 구조화된 경우 조문번호, 아니면 "조문" 컬럼
                        if "조문번호" in row.index and pd.notna(row.get("조문번호")):
                            aid = _article_num_display(str(row['조문번호']))
                        else:
                            aid = str(row.get("조문", ""))

                        orig = _clean_text(str(row.get("원문", ""))) if pd.notna(row.get("원문")) else ""
                        gem = _clean_text(_clean_translation_output(str(row.get("Gemini 번역", "")))) if pd.notna(row.get("Gemini 번역")) else ""
                        cla = _clean_text(_clean_translation_output(str(row.get("Claude 번역", "")))) if pd.notna(row.get("Claude 번역")) else ""

                        full_original.append(f"[{aid}]\n{orig}")
                        full_gemini.append(f"[{aid}]\n{gem}")
                        full_claude.append(f"[{aid}]\n{cla}")

                        # 한국법 매칭 정보
                        korean_info = ""
                        if has_matching:
                            similar_korean = str(row.get("유사 한국법", "")) if pd.notna(row.get("유사 한국법")) else ""

                            if similar_korean and similar_korean != "-":
                                korean_info = f"{_esc(similar_korean)}"

                            full_korean.append(f"[{aid}] {similar_korean}")

                        table_html += f"<tr><td><strong>{_esc(aid)}</strong></td>"
                        table_html += f"<td>{_esc(orig)}</td>"
                        table_html += f"<td>{_esc(gem)}</td>"
                        table_html += f"<td>{_esc(cla)}</td>"

                        if has_matching:
                            table_html += f"<td>{korean_info}</td>"

                        table_html += "</tr>"

                    table_html += "</tbody></table>"
                    st.html(table_html)

                    # 텍스트 복사용 영역
                    st.divider()
                    st.subheader("텍스트 복사")

                    if has_matching and full_korean:
                        # 매칭 정보가 있으면 4열로 표시
                        copy_col1, copy_col2, copy_col3, copy_col4 = st.columns(4)
                        with copy_col1:
                            st.text_area("원문 전체", "\n\n".join(full_original), height=400, key="copy_orig")
                        with copy_col2:
                            st.text_area("Gemini 번역 전체", "\n\n".join(full_gemini), height=400, key="copy_gem")
                        with copy_col3:
                            st.text_area("Claude 번역 전체", "\n\n".join(full_claude), height=400, key="copy_claude")
                        with copy_col4:
                            st.text_area("유사 한국법 전체", "\n\n".join(full_korean), height=400, key="copy_korean")
                    else:
                        # 매칭 정보가 없으면 3열로 표시
                        copy_col1, copy_col2, copy_col3 = st.columns(3)
                        with copy_col1:
                            st.text_area("원문 전체", "\n\n".join(full_original), height=400, key="copy_orig")
                        with copy_col2:
                            st.text_area("Gemini 번역 전체", "\n\n".join(full_gemini), height=400, key="copy_gem")
                        with copy_col3:
                            st.text_area("Claude 번역 전체", "\n\n".join(full_claude), height=400, key="copy_claude")

                else:
                    # ── 조문별 상세 보기 ──
                    st.subheader("조문별 상세 보기")

                    has_structured = "조문번호" in df_filtered.columns

                    for idx, row in df_filtered.iterrows():
                        country_name = row.get("국가", "") if "국가" in row.index else ""

                        # 구조 정보 구성
                        structure_info = []
                        if has_structured:
                            if country_name:
                                structure_info.append(country_name)
                            elif foreign_law_name:
                                structure_info.append(foreign_law_name)
                            if "편" in row.index and pd.notna(row.get("편")) and str(row["편"]).strip():
                                structure_info.append(str(row["편"]))
                            if "장" in row.index and pd.notna(row.get("장")) and str(row["장"]).strip():
                                structure_info.append(str(row["장"]))
                            if "절" in row.index and pd.notna(row.get("절")) and str(row["절"]).strip():
                                structure_info.append(str(row["절"]))

                            # 조문번호 표시: Section/Article/§ 접두어 제거하고 숫자만 표시
                            article_num = str(row.get('조문번호', ''))
                            art_label = _article_num_display(article_num)

                            if "조문제목" in row.index and pd.notna(row.get("조문제목")) and str(row["조문제목"]).strip():
                                art_label += f" {row['조문제목']}"
                            structure_info.append(art_label)
                        else:
                            article_id = row.get("조문", f"행 {idx}")
                            if country_name:
                                structure_info.append(f"{country_name} {foreign_law_name} — {article_id}")
                            else:
                                structure_info.append(f"{foreign_law_name} — {article_id}")

                        structure_text = " — ".join(structure_info)
                        original_text = _esc(_clean_text(str(row["원문"]))) if "원문" in row.index and pd.notna(row["원문"]) else ""
                        gemini_text = _esc(_clean_text(_clean_translation_output(str(row["Gemini 번역"])))) if "Gemini 번역" in row.index and pd.notna(row.get("Gemini 번역")) else ""
                        claude_text = _esc(_clean_text(_clean_translation_output(str(row["Claude 번역"])))) if "Claude 번역" in row.index and pd.notna(row.get("Claude 번역")) else ""

                        # 구조 정보를 위에 표시
                        st.markdown(f'<div class="article-title">{structure_text}</div>', unsafe_allow_html=True)
                        st.markdown(f"""
                        <div class="article-row">
                            <div class="article-col col-original"><div class="article-col-header">원문</div><div class="article-col-body">{original_text}</div></div>
                            <div class="article-col col-gemini"><div class="article-col-header">Gemini 번역</div><div class="article-col-body">{gemini_text}</div></div>
                            <div class="article-col col-claude"><div class="article-col-header">Claude 번역</div><div class="article-col-body">{claude_text}</div></div>
                        </div>
                        """, unsafe_allow_html=True)

                        # 유사 한국법
                        if "유사 한국법" in row.index and pd.notna(row.get("유사 한국법")) and str(row["유사 한국법"]).strip():
                            korean_article = str(row["유사 한국법"])
                            if not korean_article.startswith("["):
                                korean_article = f"[한국법] {korean_article}"
                            score_str = f" (유사도: {row['매칭 점수']})" if "매칭 점수" in row.index and pd.notna(row.get("매칭 점수")) else ""
                            korean_text_html = ""
                            if "한국법 조문 내용" in row.index and pd.notna(row.get("한국법 조문 내용")) and str(row["한국법 조문 내용"]).strip():
                                korean_text_html = f"<br><div style='margin-top:6px;white-space:pre-wrap;color:#333;'>{_esc(_clean_text(str(row['한국법 조문 내용'])))}</div>"
                            st.markdown(f'<div class="korea-law-box"><strong>유사 한국법: {korean_article}{score_str}</strong>{korean_text_html}</div>', unsafe_allow_html=True)

                            # 매칭 이유
                            if "매칭 이유" in row.index and pd.notna(row.get("매칭 이유")) and str(row["매칭 이유"]).strip():
                                st.markdown(f'<div class="diff-box"><strong>매칭 이유</strong><br>{_esc(str(row["매칭 이유"]))}</div>', unsafe_allow_html=True)

                        st.markdown("<br>", unsafe_allow_html=True)
                        st.divider()

                # Excel 다운로드
                st.divider()
                download_base = csv_name

                excel_buf = io.BytesIO()
                with pd.ExcelWriter(excel_buf, engine="openpyxl") as ew:
                    df_filtered.to_excel(ew, index=False, sheet_name="번역결과")

                    # 셀 포맷팅: 줄바꿈 표시
                    worksheet = ew.sheets["번역결과"]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                excel_download = excel_buf.getvalue()

                st.download_button("Excel 다운로드", excel_download, f"{download_base}.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="xlsx_dl")
